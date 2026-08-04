#!/usr/bin/env python3
"""最简单的本地网页：展示日报里的“快速预测”表格，并支持一键刷新。"""

from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from urllib.parse import urlparse
from pathlib import Path
import json
import os
import subprocess
import sys
import signal
import time
from zipfile import BadZipFile
from datetime import datetime

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

BASE_DIR = Path(__file__).resolve().parent.parent
OUTPUT_DIR = BASE_DIR / "data" / "output"
SCRIPTS_DIR = BASE_DIR / "scripts"
REFRESH_TIMEOUT = 90
REFRESH_LOCK_PATH = BASE_DIR / "logs" / "minutely_refresh.lock"
WEB_HEADER_MAP = {
    "目标": "目标",
    "当前间隔(手)": "当前间隔",
    "校准概率": "综合预测概率",
    "概率等级": "概率等级",
    "历史中位间隔": "历史中位间隔",
    "5手累计": "5手累计",
}


def find_latest_report():
    candidates = sorted(OUTPUT_DIR.glob("defense_summary_*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    return candidates[0] if candidates else None


def _fmt_percent(value):
    try:
        return f"{float(value) * 100:.1f}%"
    except Exception:
        return "-"


def load_quick_prediction():
    if load_workbook is None:
        return {"error": "缺少 openpyxl，无法读取 Excel。"}

    report = find_latest_report()
    if report is None:
        return {"error": "没有找到日报文件，请先运行 daily_defense_summary.py。"}

    wb = None
    for attempt in range(3):
        try:
            wb = load_workbook(report, data_only=True)
            break
        except (BadZipFile, PermissionError, OSError):
            if attempt == 2:
                return {"error": "日报正在更新，请稍后刷新。"}
            time.sleep(0.15)
    if "快速预测" not in wb.sheetnames:
        return {"error": "日报里没有“快速预测”页。"}

    ws = wb["快速预测"]
    title = ws.cell(row=1, column=1).value or "快速预测"
    headers = []
    for c in range(1, ws.max_column + 1):
        value = ws.cell(row=3, column=c).value
        if value in (None, ""):
            break
        headers.append(value)

    visible_headers = [header for header in headers if header in WEB_HEADER_MAP]
    display_headers = [WEB_HEADER_MAP[header] for header in visible_headers]

    rows = []
    for r in range(4, ws.max_row + 1):
        first = ws.cell(row=r, column=1).value
        if first in (None, ""):
            continue
        row = {}
        for c, header in enumerate(headers, 1):
            if header not in visible_headers:
                continue
            value = ws.cell(row=r, column=c).value
            display_header = WEB_HEADER_MAP[header]
            if header in ("预测概率", "校准概率", "校准前后差值", "经验间隔概率", "间隔分位", "5手累计"):
                row[display_header] = _fmt_percent(value)
            elif header == "相对倍率":
                row[display_header] = round(float(value), 2) if value is not None else None
            else:
                row[display_header] = value
        rows.append(row)

    return {
        "title": title,
        "report": report.name,
        "updated_at": datetime.fromtimestamp(report.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
        "headers": display_headers,
        "rows": rows,
    }


def refresh_pipeline():
    try:
        lock_fd = os.open(REFRESH_LOCK_PATH, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
    except FileExistsError:
        return {"error": "系统正在进行分钟刷新，请稍后再按刷新。"}
    steps = [
        ("采集", SCRIPTS_DIR / "collect_defense_data.py"),
        ("校准", SCRIPTS_DIR / "calibrate_collected_data.py"),
        ("日报", SCRIPTS_DIR / "daily_defense_summary.py"),
    ]
    logs = []
    try:
        os.close(lock_fd)
        for label, script_path in steps:
            try:
                result = subprocess.run(
                    [sys.executable, str(script_path)],
                    cwd=str(BASE_DIR), capture_output=True, text=True, timeout=REFRESH_TIMEOUT,
                )
            except subprocess.TimeoutExpired:
                return {"error": f"{label}超时，请稍后再试。", "logs": logs}
            output = "\n".join(part for part in [result.stdout.strip(), result.stderr.strip()] if part)
            logs.append({"step": label, "ok": result.returncode == 0, "output": output[-1200:]})
            if result.returncode != 0:
                return {"error": f"{label}失败。", "logs": logs}
        payload = load_quick_prediction()
        payload["logs"] = logs
        payload["pipeline_updated_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        return payload
    finally:
        try:
            REFRESH_LOCK_PATH.unlink()
        except FileNotFoundError:
            pass


HTML = """<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>快速预测</title>
  <style>
    :root {
      --bg: #f4f0e8;
      --panel: rgba(255,255,255,0.86);
      --text: #1f2937;
      --muted: #6b7280;
      --accent: #1d4ed8;
      --border: #d6d3d1;
      --shadow: 0 12px 32px rgba(0,0,0,0.10);
    }
    * { box-sizing: border-box; }
    body {
      margin: 0;
      font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", "PingFang SC", "Hiragino Sans GB", "Microsoft YaHei", sans-serif;
      color: var(--text);
      background:
        radial-gradient(circle at top left, rgba(29,78,216,0.16), transparent 30%),
        radial-gradient(circle at top right, rgba(245,158,11,0.16), transparent 28%),
        linear-gradient(180deg, #faf7f2, #f1efe9 55%, #ece7de);
      min-height: 100vh;
    }
    .wrap { max-width: 1100px; margin: 0 auto; padding: 28px 18px 40px; }
    .hero {
      background: var(--panel);
      backdrop-filter: blur(10px);
      border: 1px solid rgba(214,211,209,0.75);
      box-shadow: var(--shadow);
      border-radius: 20px;
      padding: 24px;
      margin-bottom: 18px;
    }
    h1 { margin: 0 0 10px; font-size: 30px; letter-spacing: 0.02em; }
    .sub { color: var(--muted); line-height: 1.6; font-size: 14px; }
    .toolbar { display: flex; gap: 12px; align-items: center; margin-top: 18px; flex-wrap: wrap; }
    button {
      appearance: none;
      border: 0;
      background: var(--accent);
      color: white;
      padding: 11px 16px;
      border-radius: 12px;
      cursor: pointer;
      font-size: 14px;
      box-shadow: 0 8px 18px rgba(29,78,216,0.25);
    }
    button:disabled { opacity: 0.65; cursor: wait; }
    .auto-switch {
      display: inline-flex;
      align-items: center;
      gap: 8px;
      font-size: 14px;
      color: var(--text);
      user-select: none;
      cursor: pointer;
    }
    .auto-switch input {
      width: 16px;
      height: 16px;
      accent-color: var(--accent);
    }
    .meta { color: var(--muted); font-size: 13px; }
    .card {
      background: var(--panel);
      backdrop-filter: blur(10px);
      border: 1px solid rgba(214,211,209,0.75);
      box-shadow: var(--shadow);
      border-radius: 20px;
      overflow: hidden;
    }
    table { width: 100%; border-collapse: collapse; }
    thead th {
      background: #22304a;
      color: #fff;
      text-align: left;
      font-weight: 600;
      padding: 14px 12px;
      font-size: 14px;
      white-space: nowrap;
    }
    tbody td {
      border-top: 1px solid var(--border);
      padding: 13px 12px;
      font-size: 14px;
      white-space: nowrap;
    }
    tbody tr:hover { background: rgba(29,78,216,0.04); }
    .pill {
      display: inline-block;
      padding: 4px 9px;
      border-radius: 999px;
      background: rgba(29,78,216,0.10);
      color: #1e3a8a;
      font-weight: 600;
    }
    .err { color: #b91c1c; padding: 18px 0; }
    .footer { color: var(--muted); font-size: 12px; padding: 10px 2px 0; }
    .table-wrap { overflow-x: auto; }
  </style>
</head>
<body>
  <div class="wrap">
    <div class="hero">
      <h1 id="title">快速预测</h1>
      <div class="sub" id="subtitle">正在读取最新日报...</div>
      <div class="toolbar">
        <button id="refreshBtn">立即采集并刷新</button>
        <label class="auto-switch">
          <input id="autoRefresh" type="checkbox" />
          <span>自动每分钟05秒读取最新结果</span>
        </label>
        <div class="meta" id="status">等待加载</div>
      </div>
    </div>
    <div class="card">
      <div class="table-wrap">
        <table>
          <thead><tr id="thead"></tr></thead>
          <tbody id="tbody"></tbody>
        </table>
      </div>
    </div>
    <div class="footer">本地页面自动读取最新的 `defense_summary_*.xlsx`。</div>
  </div>
  <script>
    const btn = document.getElementById('refreshBtn');
    const autoRefresh = document.getElementById('autoRefresh');
    const title = document.getElementById('title');
    const subtitle = document.getElementById('subtitle');
    const statusEl = document.getElementById('status');
    const thead = document.getElementById('thead');
    const tbody = document.getElementById('tbody');
    let autoTimer = null;
    let autoBusy = false;
    let lastReportUpdatedAt = '';
    const AUTO_REFRESH_MAX_WAIT_MS = 12000;
    const AUTO_REFRESH_POLL_MS = 800;

    function renderTable(data) {
      thead.innerHTML = '';
      tbody.innerHTML = '';
      (data.headers || []).forEach(h => {
        const th = document.createElement('th');
        th.textContent = h || '';
        thead.appendChild(th);
      });
      (data.rows || []).forEach(row => {
        const tr = document.createElement('tr');
        (data.headers || []).forEach(h => {
          const td = document.createElement('td');
          const value = row[h];
          if (h === '概率等级' && value) {
            const span = document.createElement('span');
            span.className = 'pill';
            span.textContent = value;
            td.appendChild(span);
          } else {
            td.textContent = value === null || value === undefined ? '' : value;
          }
          tr.appendChild(td);
        });
        tbody.appendChild(tr);
      });
    }

    function clearAutoTimer() {
      if (autoTimer) {
        clearTimeout(autoTimer);
        autoTimer = null;
      }
    }

    function scheduleAutoRefresh() {
      clearAutoTimer();
      if (!autoRefresh.checked) {
        return;
      }
      const now = new Date();
      const next = new Date(now);
      next.setSeconds(5, 0);
      if (next <= now) {
        next.setMinutes(next.getMinutes() + 1);
      }
      const delay = next.getTime() - now.getTime();
      autoTimer = setTimeout(async () => {
        if (!autoRefresh.checked) {
          return;
        }
        if (autoBusy) {
          scheduleAutoRefresh();
          return;
        }
        autoBusy = true;
        try {
          await loadData(false, false, true);
        } finally {
          autoBusy = false;
          scheduleAutoRefresh();
        }
      }, delay);
      statusEl.textContent = `自动刷新已开启，下次刷新时间：${next.toLocaleTimeString('zh-CN', { hour12: false })}`;
    }

    async function fetchPayload(runPipeline = false) {
      const endpoint = runPipeline ? '/api/refresh' : '/api/quick-prediction';
      const res = await fetch(endpoint + '?ts=' + Date.now(), { cache: 'no-store' });
      return res.json();
    }

    async function loadData(reschedule = true, runPipeline = false, waitForNewReport = false) {
      btn.disabled = true;
      statusEl.textContent = runPipeline ? '采集并刷新中...' : '读取中...';
      try {
        let data = await fetchPayload(runPipeline);
        if (!runPipeline && waitForNewReport && lastReportUpdatedAt) {
          const startedAt = Date.now();
          while (!data.error && data.updated_at === lastReportUpdatedAt && Date.now() - startedAt < AUTO_REFRESH_MAX_WAIT_MS) {
            statusEl.textContent = '等待新报表写入...';
            await new Promise(resolve => setTimeout(resolve, AUTO_REFRESH_POLL_MS));
            data = await fetchPayload(false);
          }
        }
        if (data.error) throw new Error(data.error);
        title.textContent = data.title || '快速预测';
        const refreshed = data.pipeline_updated_at ? ` · 采集完成: ${data.pipeline_updated_at}` : '';
        subtitle.textContent = `来源: ${data.report} · 报表时间: ${data.updated_at}${refreshed}`;
        lastReportUpdatedAt = data.updated_at || lastReportUpdatedAt;
        statusEl.textContent = '已更新';
        renderTable(data);
      } catch (err) {
        statusEl.textContent = '加载失败';
        tbody.innerHTML = '<tr><td class="err" colspan="' + Math.max((thead.children || []).length, 1) + '">' + err.message + '</td></tr>';
      } finally {
        btn.disabled = false;
        if (reschedule && autoRefresh.checked) {
          scheduleAutoRefresh();
        }
      }
    }

    btn.addEventListener('click', () => loadData(true, true));
    autoRefresh.addEventListener('change', () => {
      if (autoRefresh.checked) {
        scheduleAutoRefresh();
        loadData();
      } else {
        clearAutoTimer();
        statusEl.textContent = '自动刷新已关闭';
      }
    });
    loadData();
  </script>
</body>
</html>
"""


class Handler(BaseHTTPRequestHandler):
    def _send(self, content, content_type="text/html; charset=utf-8", status=200):
        data = content.encode("utf-8") if isinstance(content, str) else content
        self.send_response(status)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def do_GET(self):
        parsed = urlparse(self.path)
        if parsed.path == "/":
            self._send(HTML)
            return
        if parsed.path == "/api/quick-prediction":
            payload = load_quick_prediction()
            self._send(json.dumps(payload, ensure_ascii=False), content_type="application/json; charset=utf-8")
            return
        if parsed.path == "/api/refresh":
            payload = refresh_pipeline()
            self._send(json.dumps(payload, ensure_ascii=False), content_type="application/json; charset=utf-8")
            return
        self._send("Not Found", status=404)

    def log_message(self, format, *args):
        return


class ReusableThreadingHTTPServer(ThreadingHTTPServer):
    allow_reuse_address = True


def main():
    host = os.environ.get("QUICK_PREDICTION_HOST", "127.0.0.1")
    try:
        port = int(os.environ.get("QUICK_PREDICTION_PORT", "8000"))
    except ValueError:
        port = 8000
    server = None
    last_error = None
    for candidate_port in range(port, port + 11):
        try:
            server = ReusableThreadingHTTPServer((host, candidate_port), Handler)
            port = candidate_port
            break
        except OSError as exc:
            last_error = exc
    if server is None:
        raise OSError(f"无法启动本地网页服务，已尝试端口 {port}-{port + 10}: {last_error}")
    port_file = BASE_DIR / "logs" / "quick_prediction_web.port"
    port_file.parent.mkdir(parents=True, exist_ok=True)
    port_file.write_text(str(port), encoding="utf-8")
    print(f"快速预测网页已启动: http://{host}:{port}", flush=True)
    print("按 Ctrl+C 停止")
    try:
        server.serve_forever()
    finally:
        try:
            port_file.unlink()
        except FileNotFoundError:
            pass
        server.server_close()


if __name__ == "__main__":
    main()
