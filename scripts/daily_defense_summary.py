#!/usr/bin/env python3
"""
斗鱼大话三国 - 每日汇总+概率预测+校准脚本（复用巴士项目30维融合模型框架）

7个城池：洛阳(1)、成都(2)、建业(3)、荆州(4)、长安(5)、许昌(6)、汉中(7)
上三城（稀有城池）：洛阳(1)、成都(2)、建业(3) — 类比巴士项目上三站(海岛/天空/银河)
前兆城：荆州(4) — 类比巴士项目前兆站(沙滩/冰河/雪山)
用户重点关注：洛阳和成都

功能：
  Part1: 每日汇总（各城统计、卡方检验、上三城间隔统计）
  Part2: 条件概率计算（间隔分段 + 前兆模式分组条件概率）
  Part2b: 多特征融合模型训练（特征工程+逻辑回归+特征重要性）
  Part3: 概率预测（多特征融合模型+连续概率+校准后概率）
  Part4: 概率校准管道（融合模型 Platt缩放+保序回归）
  Part5: Excel报告输出（含特征分析sheet+校准曲线sheet）

独立版本：移除 codeact_sdk 依赖，纯 Python 脚本

用法：python daily_defense_summary.py [target_date]
  target_date: YYYY-MM-DD (默认今天)
"""
import sys
import os
import json
import csv
import math
from datetime import datetime, timedelta
from collections import defaultdict, deque

# 尝试导入科学计算库
try:
    import numpy as np
    HAS_NUMPY = True
except ImportError:
    HAS_NUMPY = False

try:
    from scipy.optimize import minimize
    HAS_SCIPY = True
except ImportError:
    HAS_SCIPY = False

try:
    from sklearn.linear_model import LogisticRegression
    HAS_LOGISTIC_REGRESSION = True
except ImportError:
    HAS_LOGISTIC_REGRESSION = False

try:
    from sklearn.isotonic import IsotonicRegression
    HAS_ISOTONIC = True
except ImportError:
    HAS_ISOTONIC = False

HAS_SKLEARN = HAS_LOGISTIC_REGRESSION or HAS_ISOTONIC

HAS_NUMPY_MODEL = HAS_NUMPY


# ============================================================
# 导入集中配置
# ============================================================
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from config import (
    DATA_DIR, RAW_DIR, SUMMARY_DIR, OUTPUT_DIR,
    CSV_PATH, TOWER_CSV_PATH, GAP_COUNTER_PATH, MODEL_PARAMS_PATH, CALIBRATION_LOG_PATH,
    CALIBRATION_MODEL_PATH, MODEL_UPDATE_TEMPLATE_PATH,
    CITY_MAP, CITY_IDS, TOP3_IDS, TOP3_NAMES,
    PRECURSOR_IDS, PRECURSOR_WEIGHTS, PRECURSOR_NAMES,
    PRECURSOR_GAP_THRESHOLD, PRECURSOR_LOOKBACK, PRECURSOR_MIN_POSITIVE, PRECURSOR_MIN_TOTAL,
    STRENGTH_TIERS, THEORETICAL_PROB,
    GAP_BINS, GAP_BIN_LABELS, FINE_GAP_BINS, FINE_GAP_BIN_LABELS,
    BOLD_MODE, BOLD_FACTOR, BOLD_DYNAMIC_AMPLIFICATION, BOLD_RECALIBRATE,
    BOLD_BACKTEST_TRAIN_RATIO, BOLD_GRID_SEARCH_FACTORS,
    EWMA_ALPHA, BRIER_THRESHOLD, ISOTONIC_MIN_SAMPLES, ISOTONIC_MIN_POSITIVE,
    CALIBRATION_N_BINS, CALIBRATION_MIX_FACTOR, CALIBRATION_ADJUSTMENT_WEIGHT,
    FUSION_FEATURE_NAMES, FUSION_FEATURE_DISPLAY,
    PERIOD_DEFS, PERIOD_NAMES,
    FUSION_LOOKBACK_SHORT, FUSION_LOOKBACK_MEDIUM,
    FUSION_DAYTIME_START, FUSION_DAYTIME_END,
    FUSION_MIN_POSITIVE, FUSION_MIN_TOTAL, FUSION_L2_C, FUSION_CATEGORIES,
    ELASTIC_FEATURE_NAMES, ELASTIC_FEATURE_DISPLAY,
    ELASTIC_L2_C, ELASTIC_MIN_POSITIVE, ELASTIC_MIN_TOTAL,
    SUB_MODEL_NAMES, SUB_MODEL_DISPLAY,
    DYNAMIC_WEIGHT_WINDOW, SELF_TEST_TRAIN_RATIO, SELF_TEST_N_BINS,
    EXPERIENCE_BLEND_WEIGHT, EXPERIENCE_PRIOR_STRENGTH,
    EXPERIENCE_MIN_BANDWIDTH, EXPERIENCE_BANDWIDTH_RATIO,
    BURST_LONG_GAP_THRESHOLD, BURST_WINDOW_HANDS, BURST_MIN_SAMPLES, BURST_MIN_LIFT,
    MODEL_UPDATE_TEMPLATE_VERSION,
    FLYWHEEL_LOG_PATH, FLYWHEEL_SUMMARY_PATH, ALERT_STATE_PATH,
    ALERT_TRIGGER_LEVEL, ALERT_LOG_PATH,
)
from data_flywheel import record_snapshot, resolve_pending
from notification_utils import send_dual_channel_alert


# ============================================================
# 工具函数
# ============================================================
def ensure_dirs():
    for d in [RAW_DIR, SUMMARY_DIR, OUTPUT_DIR]:
        os.makedirs(d, exist_ok=True)


def load_csv(path):
    records = []
    if not os.path.exists(path):
        return records
    with open(path, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            try:
                cid = int(row['city_id'])
                records.append((row['date'], row['time'], cid, row['city_name']))
            except (KeyError, ValueError):
                continue
    return records


def load_json(path, default=None):
    if not os.path.exists(path):
        return default if default is not None else {}
    with open(path, 'r', encoding='utf-8') as f:
        return json.load(f)


def save_json(path, data):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def get_prob_level(p):
    if p < 0.02:
        return "极低"
    if p < 0.05:
        return "低"
    if p < 0.10:
        return "中"
    if p < 0.20:
        return "中高"
    if p < 0.35:
        return "高"
    return "极高"


def build_alert_candidates(predictions):
    items = []
    for key in [TOP3_NAMES[cid] for cid in TOP3_IDS]:
        pred = predictions.get(key)
        if not pred:
            continue
        probability = pred.get("comprehensive_probability_calibrated",
                               pred.get("probability_calibrated",
                                        pred.get("comprehensive_probability",
                                                 pred.get("probability", 0.0))))
        level = get_prob_level(probability)
        if level != ALERT_TRIGGER_LEVEL:
            continue
        items.append({
            "key": key,
            "name": pred.get("name", key),
            "gap": pred.get("gap", 0),
            "probability": float(probability),
            "level": level,
            "median_gap": float((pred.get("experience_diagnostics") or {}).get("historical_gap_median", 0.0) or 0.0),
            "cum5": 1 - (1 - float(probability)) ** 5,
        })
    return items


def should_send_alerts(candidates, state):
    if not candidates:
        return False, "no_extreme_candidate", []
    if (state or {}).get("waiting_for_hit"):
        return False, "waiting_for_top3_hit", []
    signature = "|".join(
        f"{item['key']}:{item['gap']}:{item['probability']:.4f}:{item['level']}"
        for item in sorted(candidates, key=lambda item: item["key"])
    )
    if (state or {}).get("last_signature") == signature:
        return False, signature, []
    return True, signature, candidates


def append_notification_log(event, details=None):
    os.makedirs(os.path.dirname(ALERT_LOG_PATH), exist_ok=True)
    payload = {
        "at": datetime.now().isoformat(timespec="seconds"),
        "event": event,
        **(details or {}),
    }
    try:
        with open(ALERT_LOG_PATH, "a", encoding="utf-8") as f:
            f.write(json.dumps(payload, ensure_ascii=False) + "\n")
    except OSError:
        pass


def notify_extreme_predictions(predictions):
    candidates = build_alert_candidates(predictions)
    state = load_json(ALERT_STATE_PATH, {})

    # 一次预警后锁定通知；只有通知之后真实采集到上三城命中，才解锁下一次预警。
    if state.get("waiting_for_hit"):
        alert_at = state.get("last_sent_at", "")
        hit_after_alert = False
        if alert_at:
            try:
                alert_dt = datetime.fromisoformat(alert_at)
                for date_text, time_text, city_id, _ in load_csv(CSV_PATH):
                    record_dt = datetime.strptime(
                        f"{date_text} {time_text}", "%Y-%m-%d %H:%M"
                    )
                    if city_id in TOP3_IDS and record_dt > alert_dt:
                        hit_after_alert = True
                        break
            except (TypeError, ValueError):
                pass
        if hit_after_alert:
            state["waiting_for_hit"] = False
            state["unlocked_at"] = datetime.now().isoformat(timespec="seconds")
            save_json(ALERT_STATE_PATH, state)
            append_notification_log("unlocked_after_hit", {"alert_at": alert_at})
        else:
            append_notification_log("blocked_waiting_for_hit", {"candidates": candidates})
            return {"sent": False, "reason": "waiting_for_top3_hit", "items": candidates}

    should_send, signature, alert_items = should_send_alerts(candidates, state)
    if not should_send:
        reason = "no_extreme_candidate" if not candidates else "duplicate_extreme"
        append_notification_log(reason, {"candidates": candidates, "signature": signature})
        return {"sent": False, "reason": reason, "items": candidates}

    title = f"上三城极高预警 {datetime.now().strftime('%H:%M:%S')}"
    # 触发条件仍是任一上三城达到“极高”，但正文固定展示四项，
    # 便于收到通知时同时比较整体与三座城的等级。
    display_keys = ["any_top3"] + [TOP3_NAMES[cid] for cid in TOP3_IDS]
    lines = []
    for key in display_keys:
        pred = predictions.get(key, {})
        probability = pred.get("comprehensive_probability_calibrated",
                               pred.get("probability_calibrated",
                                        pred.get("comprehensive_probability",
                                                 pred.get("probability", 0.0))))
        name = "任一上三城" if key == "any_top3" else TOP3_NAMES.get(
            next((cid for cid in TOP3_IDS if TOP3_NAMES[cid] == key), 0), key
        )
        lines.append(f"{name}：{get_prob_level(float(probability))}")
    body = "\n".join(lines)

    results = send_dual_channel_alert(title, body)
    append_notification_log("send_result", {"candidates": alert_items, "results": results})
    if any(result.get("ok") for result in results):
        save_json(ALERT_STATE_PATH, {
            "last_signature": signature,
            "last_sent_at": datetime.now().isoformat(timespec="seconds"),
            "waiting_for_hit": True,
            "items": alert_items,
        })
        return {"sent": True, "items": alert_items, "results": results}
    return {"sent": False, "reason": "all_channels_failed", "items": alert_items, "results": results}


def get_gap_bin(gap):
    for i, (lo, hi) in enumerate(GAP_BINS):
        if lo <= gap <= hi:
            return i
    return len(GAP_BINS) - 1


def chi_square_test(observed, expected):
    chi2 = 0.0
    for o, e in zip(observed, expected):
        if e > 0:
            chi2 += (o - e) ** 2 / e
    df = len(observed) - 1
    p_value = _chi2_sf(chi2, df)
    return chi2, p_value


def _chi2_sf(x, df):
    if x <= 0:
        return 1.0
    if df == 1:
        return 2 * (1 - _normal_cdf(math.sqrt(x)))
    try:
        return _gamma_reg(df / 2, x / 2)
    except Exception:
        return math.exp(-x / 2)


def _normal_cdf(x):
    return 0.5 * (1 + math.erf(x / math.sqrt(2)))


def _gamma_reg(a, x):
    if x < a + 1:
        s = 1.0 / a
        t = 1.0 / a
        for n in range(1, 200):
            t *= x / (a + n)
            s += t
            if abs(t) < abs(s) * 1e-10:
                break
        return 1 - s * math.exp(-x + a * math.log(x) - _log_gamma(a))
    else:
        b = x + 1 - a
        c = 1e30
        d = 1 / b
        h = d
        for i in range(1, 200):
            an = -i * (i - a)
            b += 2
            d = an * d + b
            if abs(d) < 1e-30:
                d = 1e-30
            c = b + an / c
            if abs(c) < 1e-30:
                c = 1e-30
            d = 1 / d
            delta = d * c
            h *= delta
            if abs(delta - 1) < 1e-10:
                break
        return math.exp(-x + a * math.log(x) - _log_gamma(a)) * h


def _log_gamma(x):
    if x < 0.5:
        return math.log(math.pi / math.sin(math.pi * x)) - _log_gamma(1 - x)
    x -= 1
    g = 7
    coef = [
        0.99999999999980993, 676.5203681218851, -1259.1392167224028,
        771.32342877765313, -176.61502916214059, 12.507343278686905,
        -0.13857109526572012, 9.98436957867e-6, 1.50563273515e-7
    ]
    ag = coef[0]
    for i in range(1, g + 2):
        ag += coef[i] / (x + i)
    t = x + g + 0.5
    return 0.5 * math.log(2 * math.pi) + (x + 0.5) * math.log(t) - t + math.log(ag)


class NumpyStandardScaler:
    def fit(self, X):
        X = np.asarray(X, dtype=float)
        self.mean_ = X.mean(axis=0)
        self.scale_ = X.std(axis=0)
        self.scale_[self.scale_ == 0] = 1.0
        return self

    def transform(self, X):
        X = np.asarray(X, dtype=float)
        return (X - self.mean_) / self.scale_

    def fit_transform(self, X):
        return self.fit(X).transform(X)


class NumpyLogisticRegression:
    """轻量级逻辑回归后备实现，供没有 sklearn 的环境使用。"""

    def __init__(self, C=1.0, penalty='l2', solver='lbfgs', max_iter=1000, class_weight=None, lr=0.1):
        self.C = C
        self.max_iter = max_iter
        self.class_weight = class_weight
        self.lr = lr

    def fit(self, X, y):
        X = np.asarray(X, dtype=float)
        y = np.asarray(y, dtype=float)
        n_samples, n_features = X.shape
        self.coef_ = np.zeros(n_features, dtype=float)
        self.intercept_ = np.array([0.0], dtype=float)

        if self.class_weight == 'balanced':
            pos = max(1.0, float((y == 1).sum()))
            neg = max(1.0, float((y == 0).sum()))
            w_pos = neg / (pos + neg)
            w_neg = pos / (pos + neg)
        else:
            w_pos = w_neg = 1.0

        # L2 惩罚按样本数缩放，避免在 numpy 后备模型中压扁全部系数。
        reg = 1.0 / max(self.C * n_samples, 1e-6)
        for _ in range(self.max_iter):
            z = X @ self.coef_ + self.intercept_[0]
            p = 1.0 / (1.0 + np.exp(-np.clip(z, -35, 35)))
            sample_w = np.where(y > 0.5, w_pos, w_neg)
            error = (p - y) * sample_w
            grad_w = (X.T @ error) / n_samples + reg * self.coef_
            grad_b = error.mean()
            self.coef_ -= self.lr * grad_w
            self.intercept_[0] -= self.lr * grad_b
            if np.linalg.norm(grad_w) < 1e-6 and abs(grad_b) < 1e-6:
                break
        return self

    def predict_proba(self, X):
        X = np.asarray(X, dtype=float)
        z = X @ self.coef_ + self.intercept_[0]
        p = 1.0 / (1.0 + np.exp(-np.clip(z, -35, 35)))
        return np.column_stack([1.0 - p, p])

    def score(self, X, y):
        y = np.asarray(y, dtype=float)
        pred = (self.predict_proba(X)[:, 1] >= 0.5).astype(float)
        return float((pred == y).mean())


def parse_hour(time_str):
    try:
        parts = time_str.split(':')
        return int(parts[0])
    except (ValueError, IndexError):
        return 12


def hour_sin_cos(hour):
    theta = 2 * math.pi * hour / 24.0
    return math.sin(theta), math.cos(theta)


# ============================================================
# 前兆模式工具函数
# ============================================================
def get_precursor_cities_in_window(recent_cids, lookback=PRECURSOR_LOOKBACK):
    if len(recent_cids) >= lookback:
        recent = list(recent_cids)[-lookback:]
    else:
        recent = list(recent_cids)
    return [cid for cid in recent if cid in PRECURSOR_IDS]


def compute_precursor_strength(precursor_cids):
    return sum(PRECURSOR_WEIGHTS.get(cid, 0) for cid in precursor_cids)


def get_strength_tier(strength):
    if strength <= 0:
        return "无"
    elif strength <= 2:
        return "弱"
    elif strength <= 4:
        return "中"
    else:
        return "强"


# ============================================================
# Part1: 每日汇总
# ============================================================
def compute_daily_summary(records, target_date=None):
    if not records:
        return {}

    sorted_records = sorted(records, key=lambda r: (r[0], r[1]))

    # 按天分组统计
    daily_stats = defaultdict(lambda: {"total": 0, "by_city": defaultdict(int)})
    for date_str, time_str, cid, cname in sorted_records:
        daily_stats[date_str]["total"] += 1
        daily_stats[date_str]["by_city"][cid] += 1

    # 计算上三城间隔统计
    gap_sequences = {cid: [] for cid in TOP3_IDS}
    last_top3_pos = {cid: None for cid in TOP3_IDS}
    any_top3_last_pos = None
    any_top3_gaps = []

    for i, (date_str, time_str, cid, cname) in enumerate(sorted_records):
        if cid in TOP3_IDS:
            if last_top3_pos[cid] is not None:
                gap = i - last_top3_pos[cid]
                gap_sequences[cid].append(gap)
            last_top3_pos[cid] = i

            if any_top3_last_pos is not None:
                gap = i - any_top3_last_pos
                any_top3_gaps.append(gap)
            any_top3_last_pos = i

    # 各城池统计
    city_stats = {}
    for cid in CITY_IDS:
        total_attacks = sum(d["by_city"].get(cid, 0) for d in daily_stats.values())
        total_hands = sum(d["total"] for d in daily_stats.values())
        rate = total_attacks / total_hands if total_hands > 0 else 0
        city_stats[cid] = {
            "name": CITY_MAP[cid],
            "total_attacks": total_attacks,
            "rate": rate,
            "theoretical_prob": THEORETICAL_PROB.get(cid, 0),
        }

    # 卡方检验
    observed = [city_stats[cid]["total_attacks"] for cid in CITY_IDS]
    total = sum(observed)
    expected = [total * THEORETICAL_PROB.get(cid, 1/7) for cid in CITY_IDS]
    chi2, p_value = chi_square_test(observed, expected)

    # 上三城间隔统计
    top3_gap_stats = {}
    for cid in TOP3_IDS:
        gaps = gap_sequences[cid]
        if gaps:
            top3_gap_stats[cid] = {
                "name": TOP3_NAMES[cid],
                "count": len(gaps),
                "mean": sum(gaps) / len(gaps),
                "min": min(gaps),
                "max": max(gaps),
                "median": sorted(gaps)[len(gaps)//2],
            }
        else:
            top3_gap_stats[cid] = {"name": TOP3_NAMES[cid], "count": 0, "mean": 0, "min": 0, "max": 0, "median": 0}

    any_top3_gap_stat = {}
    if any_top3_gaps:
        any_top3_gap_stat = {
            "count": len(any_top3_gaps),
            "mean": sum(any_top3_gaps) / len(any_top3_gaps),
            "min": min(any_top3_gaps),
            "max": max(any_top3_gaps),
            "median": sorted(any_top3_gaps)[len(any_top3_gaps)//2],
        }

    return {
        "total_records": len(sorted_records),
        "daily_stats": dict(daily_stats),
        "city_stats": city_stats,
        "chi_square": {"chi2": chi2, "p_value": p_value},
        "top3_gap_stats": top3_gap_stats,
        "any_top3_gap_stats": any_top3_gap_stat,
        "gap_sequences": gap_sequences,
        "any_top3_gaps": any_top3_gaps,
    }


# ============================================================
# Part2: 条件概率计算
# ============================================================
def compute_conditional_probabilities(records):
    if not records:
        return {}

    sorted_records = sorted(records, key=lambda r: (r[0], r[1]))

    # 按间隔分段统计上三城出现概率
    gap_bin_stats = {label: {"total": 0, "top3_hit": 0, "by_city": defaultdict(int)} for label in GAP_BIN_LABELS}

    any_top3_last_pos = None
    for i, (date_str, time_str, cid, cname) in enumerate(sorted_records):
        if any_top3_last_pos is not None:
            gap = i - any_top3_last_pos
            bin_idx = get_gap_bin(gap)
            label = GAP_BIN_LABELS[bin_idx]
            gap_bin_stats[label]["total"] += 1
            if cid in TOP3_IDS:
                gap_bin_stats[label]["top3_hit"] += 1
                gap_bin_stats[label]["by_city"][cid] += 1

        if cid in TOP3_IDS:
            any_top3_last_pos = i

    # 计算条件概率
    cond_probs = {}
    for label in GAP_BIN_LABELS:
        s = gap_bin_stats[label]
        total = s["total"]
        top3_rate = s["top3_hit"] / total if total > 0 else 0
        city_rates = {}
        for cid in TOP3_IDS:
            city_rates[cid] = {
                "name": TOP3_NAMES[cid],
                "count": s["by_city"].get(cid, 0),
                "rate": s["by_city"].get(cid, 0) / total if total > 0 else 0,
            }
        cond_probs[label] = {
            "total": total,
            "top3_hit": s["top3_hit"],
            "top3_rate": top3_rate,
            "by_city": city_rates,
        }

    return cond_probs


# ============================================================
# Part2b: 多特征融合模型训练
# ============================================================
def compute_fusion_features(records, idx, recent_sids, current_gaps, prev_top3_id=None):
    """计算第idx条记录的6维融合特征"""
    date_str, time_str, cid, cname = records[idx]
    hour = parse_hour(time_str)
    h_sin, h_cos = hour_sin_cos(hour)

    # f1: 当前间隔手数（任一上三城）
    gap = current_gaps.get("any_top3", 0)

    # f2: 前兆强度分
    precursor_cids = get_precursor_cities_in_window(recent_sids, PRECURSOR_LOOKBACK)
    precursor_strength = compute_precursor_strength(precursor_cids)

    # f3: 近10手上三城次数
    lookback_10 = list(recent_sids)[-FUSION_LOOKBACK_SHORT:] if len(recent_sids) >= FUSION_LOOKBACK_SHORT else list(recent_sids)
    top3_count_10 = sum(1 for s in lookback_10 if s in TOP3_IDS)

    # f4: 近30手上三城次数
    lookback_30 = list(recent_sids)[-FUSION_LOOKBACK_MEDIUM:] if len(recent_sids) >= FUSION_LOOKBACK_MEDIUM else list(recent_sids)
    top3_count_30 = sum(1 for s in lookback_30 if s in TOP3_IDS)

    return {
        "gap_hands": gap,
        "precursor_strength": precursor_strength,
        "top3_count_10": top3_count_10,
        "top3_count_30": top3_count_30,
        "hour_sin": h_sin,
        "hour_cos": h_cos,
    }


def compute_elastic_features(base_features, hour, prev_top3_id=None):
    """从6维基线特征扩展到30维弹性特征"""
    gap = base_features["gap_hands"]
    precursor = base_features["precursor_strength"]
    freq10 = base_features["top3_count_10"] / FUSION_LOOKBACK_SHORT if FUSION_LOOKBACK_SHORT > 0 else 0

    # 时段指示变量
    period_morning = 1 if 9 <= hour < 12 else 0
    period_afternoon = 1 if 14 <= hour < 18 else 0
    period_evening = 1 if 18 <= hour < 21 else 0
    period_night = 1 if hour >= 21 or hour < 6 else 0

    # 马尔可夫特征
    prev_luoyang = 1 if prev_top3_id == 1 else 0
    prev_chengdu = 1 if prev_top3_id == 2 else 0
    prev_jianye = 1 if prev_top3_id == 3 else 0

    # 长间隔加成
    gap_long_bonus = gap if gap >= 50 else 0

    features = dict(base_features)
    features.update({
        "gap_sqrt": math.sqrt(gap) if gap >= 0 else 0,
        "gap_squared": gap * gap / 1000.0,  # 缩放防溢出
        "gap_x_period_morning": gap * period_morning,
        "gap_x_period_afternoon": gap * period_afternoon,
        "gap_x_period_evening": gap * period_evening,
        "gap_x_period_night": gap * period_night,
        "precursor_x_period_morning": precursor * period_morning,
        "precursor_x_period_afternoon": precursor * period_afternoon,
        "precursor_x_period_evening": precursor * period_evening,
        "precursor_x_period_night": precursor * period_night,
        "freq10_x_period_morning": freq10 * period_morning,
        "freq10_x_period_afternoon": freq10 * period_afternoon,
        "freq10_x_period_evening": freq10 * period_evening,
        "freq10_x_period_night": freq10 * period_night,
        "prev_top3_is_luoyang": prev_luoyang,
        "prev_top3_is_chengdu": prev_chengdu,
        "prev_top3_is_j0ianye": prev_jianye,
        "gap_long_bonus": gap_long_bonus,
        "prev_luoyang_x_gap": prev_luoyang * gap,
        "prev_chengdu_x_gap": prev_chengdu * gap,
        "prev_jianye_x_gap": prev_jianye * gap,
        "prev_luoyang_x_precursor": prev_luoyang * precursor,
        "prev_chengdu_x_precursor": prev_chengdu * precursor,
        "prev_jianye_x_precursor": prev_jianye * precursor,
    })
    # 修正 typo key
    features["prev_top3_is_jianye"] = features.pop("prev_top3_is_j0ianye", prev_jianye)
    return features


def train_fusion_model(records, category="any_top3", feature_names=None, min_positive=FUSION_MIN_POSITIVE, min_total=FUSION_MIN_TOTAL, l2_c=FUSION_L2_C):
    """训练多特征融合逻辑回归模型"""
    if not HAS_NUMPY:
        return None
    if feature_names is None:
        feature_names = FUSION_FEATURE_NAMES

    sorted_records = sorted(records, key=lambda r: (r[0], r[1]))
    n = len(sorted_records)
    if n < min_total:
        return None

    # 构建训练数据
    X_list = []
    y_list = []
    recent_sids = deque(maxlen=FUSION_LOOKBACK_MEDIUM)
    any_top3_last_pos = None
    target_last_pos = None
    prev_top3_id = None

    # 判断目标类别
    if category == "any_top3":
        target_ids = TOP3_IDS
    else:
        target_ids = [k for k, v in TOP3_NAMES.items() if v == category]
        if not target_ids:
            return None

    for i in range(n):
        date_str, time_str, cid, cname = sorted_records[i]
        is_target = cid in target_ids

        # 计算当前间隔
        current_gaps = {}
        current_gaps["any_top3"] = i - 1 - any_top3_last_pos if any_top3_last_pos is not None else i
        target_gap = i - 1 - target_last_pos if target_last_pos is not None else i

        # 需要有足够的历史才训练
        if i >= FUSION_LOOKBACK_MEDIUM:
            base_features = compute_fusion_features(sorted_records, i, recent_sids, current_gaps, prev_top3_id)
            # 单城模型必须使用该城自己的等待间隔，不能复用任一上三城间隔。
            base_features["gap_hands"] = target_gap

            if feature_names == ELASTIC_FEATURE_NAMES:
                hour = parse_hour(time_str)
                features = compute_elastic_features(base_features, hour, prev_top3_id)
            else:
                features = base_features

            x_row = [features.get(fn, 0) for fn in feature_names]
            X_list.append(x_row)
            y_list.append(1 if is_target else 0)

        # 更新状态
        recent_sids.append(cid)
        if cid in TOP3_IDS:
            any_top3_last_pos = i
            prev_top3_id = cid
        if is_target:
            target_last_pos = i

    X = np.array(X_list)
    y = np.array(y_list)

    positive_count = np.sum(y == 1)
    if positive_count < min_positive or len(y) < min_total:
        return None

    try:
        scaler = NumpyStandardScaler()
        X_scaled = scaler.fit_transform(X)

        if HAS_SKLEARN:
            model = LogisticRegression(
                C=l2_c, penalty='l2', solver='lbfgs', max_iter=1000,
                # 概率预测保留真实基准率；类别加权会把稀有事件概率推向50%。
                class_weight=None
            )
        else:
            model = NumpyLogisticRegression(
                C=l2_c,
                max_iter=1500,
                class_weight=None,
                lr=0.08,
            )
        model.fit(X_scaled, y)

        # 特征重要性
        coefficients = model.coef_[0] if hasattr(model.coef_, "ndim") and getattr(model.coef_, "ndim", 1) > 1 else model.coef_
        feature_importance = {}
        for fn, coef in zip(feature_names, coefficients):
            feature_importance[fn] = {
                "coefficient": float(coef),
                "abs_coefficient": float(abs(coef)),
                "display_name": ELASTIC_FEATURE_DISPLAY.get(fn, FUSION_FEATURE_DISPLAY.get(fn, fn)),
            }

        return {
            "model": model,
            "scaler": scaler,
            "feature_names": feature_names,
            "feature_importance": feature_importance,
            "intercept": float(model.intercept_[0]),
            "n_samples": len(y),
            "n_positive": int(positive_count),
            "n_negative": int(np.sum(y == 0)),
            "train_accuracy": float(model.score(X_scaled, y)),
        }
    except Exception as e:
        print(f"融合模型训练失败({category}): {e}")
        return None


# ============================================================
# Part3: 概率预测
# ============================================================
def predict_with_model(model_result, features):
    """使用融合模型预测概率"""
    if model_result is None:
        return None

    model = model_result["model"]
    scaler = model_result["scaler"]
    feature_names = model_result["feature_names"]

    x_row = np.array([[features.get(fn, 0) for fn in feature_names]])
    x_scaled = scaler.transform(x_row)
    prob = model.predict_proba(x_scaled)[0][1]
    return float(prob)


def compute_features_at_index(records, idx, feature_names, category="any_top3"):
    """只用 idx 之前的记录构造预测特征，避免把当前结果泄漏进预测。"""
    sorted_records = sorted(records, key=lambda r: (r[0], r[1]))
    history = sorted_records[:idx]
    if not history:
        return None
    recent_sids = deque([r[2] for r in history[-FUSION_LOOKBACK_MEDIUM:]], maxlen=FUSION_LOOKBACK_MEDIUM)
    last_top3_id = next((r[2] for r in reversed(history) if r[2] in TOP3_IDS), None)
    current_gaps = {}
    for cid in TOP3_IDS:
        last_pos = next((j for j in range(len(history) - 1, -1, -1) if history[j][2] == cid), None)
        current_gaps[cid] = len(history) - 1 - last_pos if last_pos is not None else len(history)
    current_gaps["any_top3"] = min(current_gaps.values())
    base = compute_fusion_features(sorted_records, idx, recent_sids, current_gaps, last_top3_id)
    if category != "any_top3":
        target_id = next((cid for cid, name in TOP3_NAMES.items() if name == category), None)
        if target_id is not None:
            base["gap_hands"] = current_gaps[target_id]
    if feature_names == ELASTIC_FEATURE_NAMES:
        return compute_elastic_features(base, parse_hour(sorted_records[idx][1]), last_top3_id)
    return base


def build_interval_experience_model(records, category="any_top3"):
    """从全部历史构建按等待间隔预测下一手的经验模型。"""
    target_ids = set(TOP3_IDS) if category == "any_top3" else {
        cid for cid, name in TOP3_NAMES.items() if name == category
    }
    if not target_ids:
        return None
    sorted_records = sorted(records, key=lambda r: (r[0], r[1]))
    samples = []
    completed_intervals = []
    last_hit = None
    for idx, record in enumerate(sorted_records):
        gap_before = idx if last_hit is None else idx - 1 - last_hit
        outcome = 1 if record[2] in target_ids else 0
        if last_hit is not None:
            samples.append((gap_before, outcome))
        if outcome:
            if last_hit is not None:
                completed_intervals.append(idx - last_hit - 1)
            last_hit = idx
    if not samples:
        return None
    gaps = sorted(completed_intervals)
    def percentile(q):
        if not gaps:
            return 0.0
        pos = (len(gaps) - 1) * q
        lo = int(math.floor(pos))
        hi = int(math.ceil(pos))
        if lo == hi:
            return float(gaps[lo])
        return float(gaps[lo] + (gaps[hi] - gaps[lo]) * (pos - lo))
    return {
        "category": category,
        "samples": samples,
        "n_samples": len(samples),
        "n_positive": int(sum(y for _, y in samples)),
        "base_rate": sum(y for _, y in samples) / len(samples),
        "completed_intervals": len(gaps),
        "gap_mean": sum(gaps) / len(gaps) if gaps else 0.0,
        "gap_median": percentile(0.5),
        "gap_p25": percentile(0.25),
        "gap_p75": percentile(0.75),
        "gap_p90": percentile(0.90),
    }


def predict_interval_experience(model, current_gap):
    """相似间隔核加权概率；返回概率和可解释诊断。"""
    if not model:
        return None, {}
    bandwidth = max(EXPERIENCE_MIN_BANDWIDTH, (current_gap + 1) * EXPERIENCE_BANDWIDTH_RATIO)
    weighted_hits = 0.0
    weight_total = 0.0
    effective = 0
    for historical_gap, outcome in model["samples"]:
        distance = (historical_gap - current_gap) / bandwidth
        weight = math.exp(-0.5 * distance * distance)
        if weight >= 0.05:
            effective += 1
        weighted_hits += weight * outcome
        weight_total += weight
    prior_hits = EXPERIENCE_PRIOR_STRENGTH * model["base_rate"]
    probability = (weighted_hits + prior_hits) / (weight_total + EXPERIENCE_PRIOR_STRENGTH)
    percentile = sum(1 for gap, _ in model["samples"] if gap <= current_gap) / model["n_samples"]
    return _clip_probability(probability), {
        "bandwidth": bandwidth,
        "similar_samples": effective,
        "gap_percentile": percentile,
        "historical_gap_mean": model["gap_mean"],
        "historical_gap_median": model["gap_median"],
        "historical_gap_p25": model["gap_p25"],
        "historical_gap_p75": model["gap_p75"],
        "historical_gap_p90": model["gap_p90"],
        "experience_probability": probability,
    }


def build_burst_pattern_model(records, category):
    """Learn whether a long completed gap is followed by a short repeat gap."""
    target_ids = set(TOP3_IDS) if category == "any_top3" else {
        cid for cid, name in TOP3_NAMES.items() if name == category
    }
    sorted_records = sorted(records, key=lambda r: (r[0], r[1]))
    hit_positions = [i for i, row in enumerate(sorted_records) if row[2] in target_ids]
    gaps = [hit_positions[i] - hit_positions[i - 1] - 1 for i in range(1, len(hit_positions))]
    if len(gaps) < 2:
        return {"category": category, "samples": 0, "eligible": False}
    base_rate = sum(g <= BURST_WINDOW_HANDS for g in gaps[1:]) / max(len(gaps) - 1, 1)
    conditioned = [gaps[i + 1] for i, gap in enumerate(gaps[:-1]) if gap >= BURST_LONG_GAP_THRESHOLD]
    burst_rate = sum(g <= BURST_WINDOW_HANDS for g in conditioned) / len(conditioned) if conditioned else 0.0
    lift = burst_rate / base_rate if base_rate > 0 else 0.0
    return {
        "category": category,
        "long_gap_threshold": BURST_LONG_GAP_THRESHOLD,
        "window_hands": BURST_WINDOW_HANDS,
        "samples": len(conditioned),
        "hits": sum(g <= BURST_WINDOW_HANDS for g in conditioned),
        "base_rate": base_rate,
        "burst_rate": burst_rate,
        "lift": lift,
        "eligible": len(conditioned) >= BURST_MIN_SAMPLES and lift >= BURST_MIN_LIFT,
    }


def predict_burst_pattern(model, current_gap, previous_gap=None):
    """Return a bounded boost and diagnostics for the long-gap burst pattern."""
    if not model or current_gap is None:
        return 0.0, {"burst_pattern": False, "burst_eligible": False}
    # The pattern is only active after the long gap has already produced one hit.
    active = (
        int(current_gap) <= int(model.get("window_hands", BURST_WINDOW_HANDS))
        and previous_gap is not None
        and int(previous_gap) >= int(model.get("long_gap_threshold", BURST_LONG_GAP_THRESHOLD))
    )
    eligible = bool(model.get("eligible")) and active
    boost = 0.0
    if eligible:
        boost = min(0.08, max(0.0, float(model.get("burst_rate", 0.0)) - float(model.get("base_rate", 0.0))) * 0.35)
    return boost, {
        "burst_pattern": active,
        "burst_previous_gap": previous_gap,
        "burst_eligible": eligible,
        "burst_samples": int(model.get("samples", 0)),
        "burst_rate": float(model.get("burst_rate", 0.0)),
        "burst_lift": float(model.get("lift", 0.0)),
        "burst_boost": boost,
    }


def blend_model_and_experience(model_probability, experience_probability, weight=None):
    if experience_probability is None:
        return _clip_probability(model_probability)
    if model_probability is None:
        return _clip_probability(experience_probability)
    weight = EXPERIENCE_BLEND_WEIGHT if weight is None else max(0.45, min(0.80, float(weight)))
    return _clip_probability((1.0 - weight) * model_probability + weight * experience_probability)


def compute_interval_pressure(current_gap, diagnostics, base_rate):
    """把“当前间隔已经偏离历史常态多少”转成额外抬升信号。"""
    if current_gap is None or not diagnostics:
        return 0.0
    median_gap = float(diagnostics.get("historical_gap_median", 0.0) or 0.0)
    p75_gap = float(diagnostics.get("historical_gap_p75", median_gap) or median_gap)
    p90_gap = float(diagnostics.get("historical_gap_p90", p75_gap) or p75_gap)
    percentile = float(diagnostics.get("gap_percentile", 0.0) or 0.0)
    if median_gap <= 0:
        return 0.0

    ratio = current_gap / max(median_gap, 1.0)
    pressure = 0.0
    if ratio > 1.0:
        pressure += min(0.08, (ratio - 1.0) * 0.035)
    if current_gap >= p75_gap > 0:
        pressure += min(0.05, max(0.0, percentile - 0.75) * 0.20)
    if current_gap >= p90_gap > 0:
        pressure += min(0.08, max(0.0, percentile - 0.90) * 0.60 + 0.03)

    # 长间隔时，至少给到比基础命中率更明显的提升空间。
    if ratio >= 2.0:
        pressure = max(pressure, min(0.18, base_rate * 1.25))
    return max(0.0, pressure)


def compute_comprehensive_probability(key, gap, model_probability, experience_probability, diagnostics, flywheel_weights, flywheel_state, burst_boost=0.0, burst_diagnostics=None):
    """综合模型、间隔经验、长间隔压力和飞轮误差，得到当前展示用总体概率。"""
    if model_probability is None and experience_probability is None:
        return None, {}

    base_rate = 0.0
    city_stats = (flywheel_state or {}).get("cities", {}).get(key, {})
    resolved_n = int(city_stats.get("n", 0) or 0)
    mean_error = float(city_stats.get("mean_error", 0.0) or 0.0)
    weight = (flywheel_weights or {}).get(key)
    blended = blend_model_and_experience(model_probability, experience_probability, weight)

    if experience_probability is not None:
        base_rate = float(experience_probability)
    elif model_probability is not None:
        base_rate = float(model_probability)

    interval_pressure = compute_interval_pressure(gap, diagnostics, base_rate)
    flywheel_bonus = 0.0
    if resolved_n >= 8 and mean_error > 0:
        flywheel_bonus = min(0.06, mean_error * 0.35)

    # 当前间隔越极端，越相信经验侧与压力侧；样本少时抬升幅度更谨慎。
    percentile = float((diagnostics or {}).get("gap_percentile", 0.0) or 0.0)
    experience_anchor = max(base_rate, blended)
    boosted = experience_anchor + interval_pressure + flywheel_bonus + float(burst_boost or 0.0)
    boosted = _clip_probability(boosted)
    confidence = min(0.82, 0.35 + percentile * 0.35 + min(resolved_n, 30) / 100.0)
    comprehensive = _clip_probability((1.0 - confidence) * blended + confidence * boosted)

    return comprehensive, {
        "blended_probability": blended,
        "interval_pressure": interval_pressure,
        "flywheel_bonus": flywheel_bonus,
        "confidence": confidence,
        "resolved_samples": resolved_n,
        "mean_error": mean_error,
        "short_gap_adjustment": 0.0,
        **(burst_diagnostics or {}),
    }


def _binary_metrics(probabilities, actuals):
    if not actuals:
        return {"n": 0, "hit_rate": 0.0, "brier": 0.0, "mean_abs_error": 0.0}
    pairs = list(zip(probabilities, actuals))
    positives = sum(actuals)
    predicted_positive = sum(p >= 0.5 for p in probabilities)
    true_positive = sum((p >= 0.5) and bool(y) for p, y in zip(probabilities, actuals))
    return {
        "n": len(pairs),
        "hit_rate": sum((p >= 0.5) == bool(y) for p, y in pairs) / len(pairs),
        "brier": sum((p - y) ** 2 for p, y in pairs) / len(pairs),
        "mean_abs_error": sum(abs(p - y) for p, y in pairs) / len(pairs),
        "base_rate": sum(actuals) / len(actuals),
        "precision": true_positive / predicted_positive if predicted_positive else 0.0,
        "recall": true_positive / positives if positives else 0.0,
    }


def run_time_series_backtest(records, train_ratio=SELF_TEST_TRAIN_RATIO):
    """留出后段历史做真实回测；每个样本的特征只读取该样本之前的结果。"""
    sorted_records = sorted(records, key=lambda r: (r[0], r[1]))
    split = max(FUSION_LOOKBACK_MEDIUM + 1, int(len(sorted_records) * train_ratio))
    if len(sorted_records) - split < 30:
        return {"error": "留出样本不足", "split": split, "total": len(sorted_records)}

    results = {cat: {"probabilities": [], "actuals": []} for cat in FUSION_CATEGORIES}
    holdout_models = {}
    train_records = sorted_records[:split]
    experience_models = {
        "any_top3": build_interval_experience_model(train_records, "any_top3"),
        **{TOP3_NAMES[cid]: build_interval_experience_model(train_records, TOP3_NAMES[cid]) for cid in TOP3_IDS},
    }
    for cat in FUSION_CATEGORIES:
        mr = train_fusion_model(train_records, category=cat, feature_names=ELASTIC_FEATURE_NAMES,
                                min_positive=ELASTIC_MIN_POSITIVE, min_total=ELASTIC_MIN_TOTAL,
                                l2_c=ELASTIC_L2_C)
        if mr is None:
            mr = train_fusion_model(train_records, category=cat, feature_names=FUSION_FEATURE_NAMES,
                                    min_positive=FUSION_MIN_POSITIVE, min_total=FUSION_MIN_TOTAL,
                                    l2_c=FUSION_L2_C)
        holdout_models[cat] = mr

    top1_hits = 0
    top1_total = 0
    top3_rank_hits = 0
    top3_rank_total = 0
    for idx in range(split, len(sorted_records)):
        features_by_model = {}
        for cat, mr in holdout_models.items():
            if mr is not None:
                features_by_model[cat] = compute_features_at_index(sorted_records, idx, mr["feature_names"], cat)
        city_probs = []
        for cat in ["洛阳", "成都", "建业"]:
            mr = holdout_models.get(cat)
            feat = features_by_model.get(cat)
            if mr is None or feat is None:
                continue
            prob = predict_with_model(mr, feat)
            target_id = next(cid for cid, name in TOP3_NAMES.items() if name == cat)
            history = sorted_records[:idx]
            target_gap = 0
            for j in range(len(history) - 1, -1, -1):
                if history[j][2] == target_id:
                    target_gap = len(history) - 1 - j
                    break
            exp_prob, _ = predict_interval_experience(experience_models.get(cat), target_gap)
            prob = blend_model_and_experience(prob, exp_prob)
            results[cat]["probabilities"].append(prob)
            results[cat]["actuals"].append(1 if sorted_records[idx][2] == next(cid for cid, name in TOP3_NAMES.items() if name == cat) else 0)
            city_probs.append((prob, cat))
        if city_probs:
            top1_total += 1
            if max(city_probs)[1] == TOP3_NAMES.get(sorted_records[idx][2], ""):
                top1_hits += 1
            if sorted_records[idx][2] in TOP3_IDS:
                top3_rank_total += 1
                if max(city_probs)[1] == TOP3_NAMES[sorted_records[idx][2]]:
                    top3_rank_hits += 1

        mr = holdout_models.get("any_top3")
        feat = features_by_model.get("any_top3")
        if mr is not None and feat is not None:
            raw_prob = predict_with_model(mr, feat)
            history = sorted_records[:idx]
            any_gap = 0
            for j in range(len(history) - 1, -1, -1):
                if history[j][2] in TOP3_IDS:
                    any_gap = len(history) - 1 - j
                    break
            exp_prob, _ = predict_interval_experience(experience_models.get("any_top3"), any_gap)
            results["any_top3"]["probabilities"].append(blend_model_and_experience(raw_prob, exp_prob))
            results["any_top3"]["actuals"].append(1 if sorted_records[idx][2] in TOP3_IDS else 0)

    metrics = {cat: _binary_metrics(v["probabilities"], v["actuals"]) for cat, v in results.items()}
    metrics["top1_city"] = {
        "n": top1_total,
        "hit_rate": top1_hits / top1_total if top1_total else 0.0,
        "top3_conditional_n": top3_rank_total,
        "top3_conditional_hit_rate": top3_rank_hits / top3_rank_total if top3_rank_total else 0.0,
    }
    metrics.update({"total": len(sorted_records), "train": split, "test": len(sorted_records) - split,
                    "train_ratio": split / len(sorted_records)})
    # 留出集上的原始概率与真实结果用于概率校准，避免使用训练集内预测造成泄漏。
    metrics["calibration_samples"] = results
    return metrics


def compute_current_predictions(records, model_results=None, gap_counter=None):
    """计算当前各上三城的预测概率"""
    if not records:
        return {}

    sorted_records = sorted(records, key=lambda r: (r[0], r[1]))
    n = len(sorted_records)

    # 优先使用 gap_counter（基于 tower 数据，间隔更准确）
    current_gaps = {}
    if gap_counter:
        for cid in TOP3_IDS:
            cid_str = str(cid)
            if cid_str in gap_counter and gap_counter[cid_str].get("gap_hands") is not None:
                current_gaps[cid] = gap_counter[cid_str]["gap_hands"]
            else:
                # fallback: 从 records 计算
                last_pos = None
                for i in range(n - 1, -1, -1):
                    if sorted_records[i][2] == cid:
                        last_pos = i
                        break
                current_gaps[cid] = n - 1 - last_pos if last_pos is not None else n
        current_gaps["any_top3"] = min(current_gaps.values())
    else:
        # 无 gap_counter 时从 records 计算
        for cid in TOP3_IDS:
            last_pos = None
            for i in range(n - 1, -1, -1):
                if sorted_records[i][2] == cid:
                    last_pos = i
                    break
            current_gaps[cid] = n - 1 - last_pos if last_pos is not None else n
        current_gaps["any_top3"] = min(current_gaps.values())

    # 最近N手的站点ID序列
    recent_sids = deque([r[2] for r in sorted_records[-FUSION_LOOKBACK_MEDIUM:]], maxlen=FUSION_LOOKBACK_MEDIUM)

    # 上一个上三城ID
    prev_top3_id = None
    for i in range(n - 1, -1, -1):
        if sorted_records[i][2] in TOP3_IDS:
            prev_top3_id = sorted_records[i][2]
            break

    # 当前时间特征
    now = datetime.now()
    hour = now.hour

    # 基线6维特征
    base_features = compute_fusion_features(sorted_records, n - 1, recent_sids, current_gaps, prev_top3_id)

    predictions = {}
    flywheel_state = load_json(FLYWHEEL_SUMMARY_PATH, {})
    flywheel_weights = flywheel_state.get("weights", {})
    experience_models = {
        "any_top3": build_interval_experience_model(sorted_records, "any_top3"),
        **{TOP3_NAMES[cid]: build_interval_experience_model(sorted_records, TOP3_NAMES[cid]) for cid in TOP3_IDS},
    }
    burst_models = {
        "any_top3": build_burst_pattern_model(sorted_records, "any_top3"),
        **{TOP3_NAMES[cid]: build_burst_pattern_model(sorted_records, TOP3_NAMES[cid]) for cid in TOP3_IDS},
    }

    def previous_completed_gap(key):
        target_ids = set(TOP3_IDS) if key == "any_top3" else {
            cid for cid, name in TOP3_NAMES.items() if name == key
        }
        hit_positions = [i for i, row in enumerate(sorted_records) if row[2] in target_ids]
        if len(hit_positions) < 2:
            return None
        return hit_positions[-1] - hit_positions[-2] - 1

    def add_experience(key, raw_model_probability, gap):
        experience_probability, diagnostics = predict_interval_experience(experience_models.get(key), gap)
        burst_boost, burst_diagnostics = predict_burst_pattern(
            burst_models.get(key), gap, previous_completed_gap(key)
        )
        diagnostics.update(burst_diagnostics)
        final_probability = blend_model_and_experience(raw_model_probability, experience_probability, flywheel_weights.get(key))
        comprehensive_probability, comprehensive_diagnostics = compute_comprehensive_probability(
            key, gap, raw_model_probability, experience_probability, diagnostics, flywheel_weights, flywheel_state,
            burst_boost, burst_diagnostics
        )
        return final_probability, experience_probability, diagnostics, comprehensive_probability, comprehensive_diagnostics

    # 任一上三城概率
    if model_results and "any_top3" in model_results and model_results["any_top3"] is not None:
        mr = model_results["any_top3"]
        features = compute_elastic_features(base_features, hour, prev_top3_id) if mr["feature_names"] == ELASTIC_FEATURE_NAMES else base_features
        model_prob = predict_with_model(mr, features)
        prob, exp_prob, diagnostics, comprehensive_probability, comprehensive_diagnostics = add_experience("any_top3", model_prob, current_gaps["any_top3"])
        predictions["any_top3"] = {
            "name": "任一上三城",
            "probability": prob,
            "comprehensive_probability": comprehensive_probability if comprehensive_probability is not None else prob,
            "gap": current_gaps["any_top3"],
            "model_probability": model_prob,
            "experience_probability": exp_prob,
            "experience_diagnostics": diagnostics,
            "comprehensive_diagnostics": comprehensive_diagnostics,
        }

    # 各上三城单独概率
    for cid in TOP3_IDS:
        cat_name = TOP3_NAMES[cid]
        if model_results and cat_name in model_results and model_results[cat_name] is not None:
            mr = model_results[cat_name]
            features = compute_elastic_features(base_features, hour, prev_top3_id) if mr["feature_names"] == ELASTIC_FEATURE_NAMES else base_features
            model_prob = predict_with_model(mr, features)
            prob, exp_prob, diagnostics, comprehensive_probability, comprehensive_diagnostics = add_experience(cat_name, model_prob, current_gaps.get(cid, 0))
            predictions[cat_name] = {
                "name": cat_name,
                "probability": prob,
                "comprehensive_probability": comprehensive_probability if comprehensive_probability is not None else prob,
                "gap": current_gaps.get(cid, 0),
                "model_probability": model_prob,
                "experience_probability": exp_prob,
                "experience_diagnostics": diagnostics,
                "comprehensive_diagnostics": comprehensive_diagnostics,
            }

    # 简单间隔模型兜底
    if not predictions:
        # 用历史频率估算
        for cid in TOP3_IDS:
            rate = THEORETICAL_PROB.get(cid, 0)
            gap = current_gaps.get(cid, 0)
            # 简单指数衰减模型
            experience_model = experience_models.get(TOP3_NAMES[cid])
            exp_prob, diagnostics = predict_interval_experience(experience_model, gap)
            prob = exp_prob if exp_prob is not None else 1 - (1 - rate) ** max(1, gap // 10 + 1)
            predictions[TOP3_NAMES[cid]] = {
                "name": TOP3_NAMES[cid],
                "probability": min(prob, 0.5),
                "comprehensive_probability": min(prob, 0.5),
                "gap": gap,
                "model_probability": None,
                "experience_probability": exp_prob,
                "experience_diagnostics": diagnostics,
                "comprehensive_diagnostics": {},
            }
        any_rate = sum(THEORETICAL_PROB.get(cid, 0) for cid in TOP3_IDS)
        any_gap = current_gaps["any_top3"]
        any_prob, any_diag = predict_interval_experience(experience_models.get("any_top3"), any_gap)
        if any_prob is None:
            any_prob = 1 - (1 - any_rate) ** max(1, any_gap // 10 + 1)
        predictions["any_top3"] = {
            "name": "任一上三城",
            "probability": min(any_prob, 0.8),
            "comprehensive_probability": min(any_prob, 0.8),
            "gap": any_gap,
            "model_probability": None,
            "experience_probability": any_prob,
            "experience_diagnostics": any_diag,
            "comprehensive_diagnostics": {},
        }

    return predictions


# ============================================================
# Part4: 概率校准
# ============================================================
def _clip_probability(value):
    return max(0.001, min(0.999, float(value)))


def _logit(value):
    p = _clip_probability(value)
    return math.log(p / (1.0 - p))


def _sigmoid(value):
    if value >= 0:
        z = math.exp(-value)
        return 1.0 / (1.0 + z)
    z = math.exp(value)
    return z / (1.0 + z)


def _fit_isotonic_pava(probabilities, actuals):
    """纯 Python 保序回归，避免日报依赖 sklearn 才能真正校准。"""
    pairs = sorted(zip(probabilities, actuals), key=lambda item: item[0])
    blocks = []
    for x, y in pairs:
        blocks.append({"x_min": x, "x_max": x, "sum": float(y), "n": 1})
        while len(blocks) >= 2:
            left, right = blocks[-2], blocks[-1]
            if left["sum"] / left["n"] <= right["sum"] / right["n"]:
                break
            merged = {
                "x_min": left["x_min"],
                "x_max": right["x_max"],
                "sum": left["sum"] + right["sum"],
                "n": left["n"] + right["n"],
            }
            blocks[-2:] = [merged]
    return {
        "method": "isotonic",
        "x_thresholds": [float(block["x_max"]) for block in blocks],
        "y_thresholds": [_clip_probability(block["sum"] / block["n"]) for block in blocks],
        "n_samples": len(actuals),
        "n_positive": int(sum(actuals)),
    }


def _fit_platt_python(probabilities, actuals):
    """用 Newton 法拟合 Platt scaling: sigmoid(intercept + coef * logit(p))。"""
    xs = [_logit(p) for p in probabilities]
    positive = sum(actuals)
    negative = len(actuals) - positive
    intercept = math.log((positive + 1.0) / (negative + 1.0))
    coef = 1.0
    for _ in range(100):
        gradient_a = 0.0
        gradient_b = 0.0
        h_aa = 1e-6
        h_ab = 0.0
        h_bb = 1e-6
        for x, y in zip(xs, actuals):
            q = _sigmoid(intercept + coef * x)
            weight = max(q * (1.0 - q), 1e-8)
            error = q - y
            gradient_a += error
            gradient_b += error * x
            h_aa += weight
            h_ab += weight * x
            h_bb += weight * x * x
        determinant = h_aa * h_bb - h_ab * h_ab
        if abs(determinant) < 1e-12:
            break
        step_a = (h_bb * gradient_a - h_ab * gradient_b) / determinant
        step_b = (-h_ab * gradient_a + h_aa * gradient_b) / determinant
        intercept -= step_a
        coef -= step_b
        if max(abs(step_a), abs(step_b)) < 1e-7:
            break
    return {
        "method": "platt",
        "coef": float(coef),
        "intercept": float(intercept),
        "n_samples": len(actuals),
        "n_positive": int(positive),
    }


def fit_probability_calibrators(backtest):
    """用严格时间留出样本拟合并保存每个目标的概率校准器。"""
    samples = (backtest or {}).get("calibration_samples", {})
    saved = {"version": 2, "source": "strict_time_series_holdout_blended_interval_experience", "models": {}}
    for category, sample in samples.items():
        probabilities = sample.get("probabilities", [])
        actuals = sample.get("actuals", [])
        if len(probabilities) < 30 or len(set(actuals)) < 2:
            continue
        probabilities = [_clip_probability(p) for p in probabilities]
        if (len(probabilities) >= ISOTONIC_MIN_SAMPLES and
                sum(actuals) >= ISOTONIC_MIN_POSITIVE and
                len(actuals) - sum(actuals) >= ISOTONIC_MIN_POSITIVE):
            if HAS_ISOTONIC:
                model = IsotonicRegression(y_min=0.001, y_max=0.999, out_of_bounds="clip")
                model.fit(probabilities, actuals)
                saved["models"][category] = {
                    "method": "isotonic",
                    "x_thresholds": [float(x) for x in model.X_thresholds_],
                    "y_thresholds": [float(y) for y in model.y_thresholds_],
                    "n_samples": len(actuals),
                    "n_positive": int(sum(actuals)),
                }
            else:
                saved["models"][category] = _fit_isotonic_pava(probabilities, actuals)
        else:
            if sum(actuals) < 10 or len(actuals) - sum(actuals) < 10:
                continue
            if HAS_LOGISTIC_REGRESSION:
                model = LogisticRegression(C=1e6, solver="lbfgs")
                model.fit([[_logit(p)] for p in probabilities], actuals)
                saved["models"][category] = {
                    "method": "platt",
                    "coef": float(model.coef_[0][0]),
                    "intercept": float(model.intercept_[0]),
                    "n_samples": len(actuals),
                    "n_positive": int(sum(actuals)),
                }
            else:
                saved["models"][category] = _fit_platt_python(probabilities, actuals)
    save_json(CALIBRATION_MODEL_PATH, saved)
    return saved


def apply_probability_calibrator(probability, model):
    """应用已持久化的 isotonic 或 Platt 校准器。"""
    p = _clip_probability(probability)
    if not model:
        return p
    if model.get("method") == "platt":
        return _clip_probability(_sigmoid(model["intercept"] + model["coef"] * _logit(p)))
    if model.get("method") == "isotonic":
        xs = model.get("x_thresholds", [])
        ys = model.get("y_thresholds", [])
        if not xs or not ys:
            return p
        if p <= xs[0]:
            return _clip_probability(ys[0])
        if p >= xs[-1]:
            return _clip_probability(ys[-1])
        for i in range(1, len(xs)):
            if p <= xs[i]:
                if xs[i] == xs[i - 1]:
                    return _clip_probability(ys[i])
                ratio = (p - xs[i - 1]) / (xs[i] - xs[i - 1])
                return _clip_probability(ys[i - 1] + ratio * (ys[i] - ys[i - 1]))
    return p


def calibrate_predictions(predictions, calibration_model=None):
    """用回测校准器有限幅度纠偏，避免稀有城小样本把概率压到边界。"""
    models = (calibration_model or {}).get("models", {})
    calibrated = {}
    for key, pred in predictions.items():
        raw = _clip_probability(pred.get("probability", 0))
        comprehensive_raw = _clip_probability(pred.get("comprehensive_probability", raw))
        fully_calibrated = apply_probability_calibrator(raw, models.get(key))
        final = _clip_probability(
            raw + CALIBRATION_ADJUSTMENT_WEIGHT * (fully_calibrated - raw)
        )
        comprehensive_fully_calibrated = apply_probability_calibrator(comprehensive_raw, models.get(key))
        diagnostics = pred.get("experience_diagnostics", {}) or {}
        percentile = float(diagnostics.get("gap_percentile", 0.0) or 0.0)
        current_gap = float(pred.get("gap", 0) or 0)
        historical_median = float(diagnostics.get("historical_gap_median", 0.0) or 0.0)
        long_gap_guard = historical_median > 0 and current_gap >= historical_median * 1.8 and percentile >= 0.9
        if long_gap_guard:
            comprehensive_adjustment_weight = min(0.18, CALIBRATION_ADJUSTMENT_WEIGHT * 0.5)
            comprehensive_final = max(
                comprehensive_raw,
                _clip_probability(
                    comprehensive_raw + comprehensive_adjustment_weight * (comprehensive_fully_calibrated - comprehensive_raw)
                ),
            )
        else:
            comprehensive_final = _clip_probability(
                comprehensive_raw + CALIBRATION_ADJUSTMENT_WEIGHT * (comprehensive_fully_calibrated - comprehensive_raw)
            )
        # 洛阳短间隔历史命中率高于当前综合预测，校准完成后做小幅平滑修正；
        # 设置上限，避免短间隔样本把概率直接推入高等级。
        short_gap_adjustment = 0.0
        if key == "洛阳" and 0 <= current_gap <= 10:
            before_short_gap = comprehensive_final
            comprehensive_final = min(0.06, max(0.04, comprehensive_final))
            short_gap_adjustment = comprehensive_final - before_short_gap
        calibrated_pred = dict(pred)
        calibrated_pred["probability"] = raw
        calibrated_pred["probability_calibrated"] = final
        calibrated_pred["fully_calibrated_probability"] = fully_calibrated
        calibrated_pred["calibration_delta"] = final - raw
        calibrated_pred["comprehensive_probability"] = comprehensive_raw
        calibrated_pred["comprehensive_probability_calibrated"] = comprehensive_final
        calibrated_pred["comprehensive_fully_calibrated_probability"] = comprehensive_fully_calibrated
        calibrated_pred["comprehensive_calibration_delta"] = comprehensive_final - comprehensive_raw
        calibrated_pred["comprehensive_diagnostics"] = dict(
            pred.get("comprehensive_diagnostics", {}) or {},
            short_gap_adjustment=short_gap_adjustment,
        )
        calibrated[key] = calibrated_pred
    return calibrated


# ============================================================
# Part5: Excel报告输出
# ============================================================
def generate_excel_report(summary, cond_probs, predictions, model_results=None, target_date=None, backtest=None, flywheel_summary=None, flywheel_items=None):
    """生成Excel报告，并包含预测飞轮回放与反思页。"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        import subprocess
        subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl", "-q"], timeout=60)
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()

    # 样式定义
    header_font = Font(bold=True, size=11)
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    title_font = Font(bold=True, size=14)
    subtitle_font = Font(bold=True, size=12)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    highlight_good = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    highlight_bad = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    date_str = target_date or datetime.now().strftime("%Y-%m-%d")

    # ========== Sheet 1: 快速预测 ==========
    ws1 = wb.active
    ws1.title = "快速预测"
    ws1.cell(row=1, column=1, value=f"斗鱼大话三国 - 快速预测 ({date_str})").font = title_font

    row = 3
    headers = ["目标", "当前间隔(手)", "预测概率", "校准概率", "校准前后差值", "经验间隔概率", "相似样本数", "间隔分位", "历史中位间隔", "概率等级", "相对倍率", "5手累计"]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=row, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    row += 1

    # 基准概率
    baseline_any = sum(THEORETICAL_PROB.get(cid, 0) for cid in TOP3_IDS)

    for key in ["any_top3"] + [TOP3_NAMES[cid] for cid in TOP3_IDS]:
        if key in predictions:
            pred = predictions[key]
            prob = pred.get("probability", 0)
            prob_cal = pred.get("probability_calibrated", prob)
            comprehensive_prob = pred.get("comprehensive_probability", prob_cal)
            comprehensive_prob_cal = pred.get("comprehensive_probability_calibrated", comprehensive_prob)
            gap = pred.get("gap", 0)
            name = pred.get("name", key)

            # 相对倍率
            if key == "any_top3":
                baseline = baseline_any
            else:
                cid = [k for k, v in TOP3_NAMES.items() if v == key]
                baseline = THEORETICAL_PROB.get(cid[0], 0) if cid else 0.01
            multiplier = comprehensive_prob_cal / baseline if baseline > 0 else 1.0

            # 累计概率: 1-(1-p)^n
            cum5 = 1 - (1 - comprehensive_prob_cal) ** 5

            level = get_prob_level(comprehensive_prob_cal)

            delta = pred.get("comprehensive_calibration_delta", comprehensive_prob_cal - comprehensive_prob)
            diagnostics = pred.get("experience_diagnostics", {})
            row_data = [name, gap, comprehensive_prob, comprehensive_prob_cal, delta,
                        pred.get("experience_probability"), diagnostics.get("similar_samples", 0),
                        diagnostics.get("gap_percentile", 0), diagnostics.get("historical_gap_median", 0),
                        level, multiplier, cum5]
            for col, val in enumerate(row_data, 1):
                cell = ws1.cell(row=row, column=col, value=val)
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
                if col in [3, 4, 5, 6, 8, 12]:
                    cell.number_format = '0.00%'
                if col == 7:
                    cell.number_format = '#,##0'
                if col == 9:
                    cell.number_format = '0.0'
                if col == 11:
                    cell.number_format = '0.00x'
                if col == 10:
                    if level in ("高", "极高"):
                        cell.fill = highlight_good
                        cell.font = Font(bold=True, color="00B050")
                    elif level in ("低", "极低"):
                        cell.fill = highlight_bad
                        cell.font = Font(bold=True, color="C00000")
            row += 1

    # ========== Sheet 8: 预测回放 ==========
    ws8 = wb.create_sheet("预测回放")
    ws8.cell(row=1, column=1, value=f"预测飞轮回放 ({date_str})").font = title_font
    replay_headers = ["快照时间", "状态", "实际记录", "城池", "预测间隔", "最终概率", "经验概率", "模型概率", "实际命中", "预测误差", "预测前间隔"]
    for col, h in enumerate(replay_headers, 1):
        cell = ws8.cell(row=3, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
    replay_row = 4
    for item in (flywheel_items or [])[-200:]:
        actual_record = item.get("actual_record", {})
        actual_text = f"{actual_record.get('date','')} {actual_record.get('time','')} {actual_record.get('city_name','')}" if actual_record else ""
        for city in ["洛阳", "成都", "建业"]:
            value = item.get("cities", {}).get(city, {})
            actual = value.get("actual")
            probability = value.get("final_probability")
            replay_values = [item.get("created_at"), item.get("status"), actual_text, city, value.get("gap"), probability, value.get("experience_probability"), value.get("model_probability"), actual, (actual - probability) if actual is not None and probability is not None else None, value.get("gap_before_prediction")]
            for col, val in enumerate(replay_values, 1):
                cell = ws8.cell(row=replay_row, column=col, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")
                if col in (6, 7, 8, 10):
                    cell.number_format = "0.00%"
            replay_row += 1

    # ========== Sheet 9: 命中反思 ==========
    ws9 = wb.create_sheet("命中反思")
    ws9.cell(row=1, column=1, value=f"命中反思与间隔经验 ({date_str})").font = title_font
    reflect_headers = ["城池", "已结算样本", "实际命中率", "平均预测概率", "平均误差(实际-预测)", "Brier", "间隔区间", "区间样本", "区间命中率", "区间平均预测"]
    for col, h in enumerate(reflect_headers, 1):
        cell = ws9.cell(row=3, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
    reflect_row = 4
    for city in ["洛阳", "成都", "建业"]:
        metric = (flywheel_summary or {}).get("cities", {}).get(city, {})
        bands = metric.get("bands", {})
        band_items = list(bands.items()) or [("暂无", {})]
        for band, b in band_items:
            values = [city, metric.get("n", 0), metric.get("hit_rate", 0), metric.get("mean_probability", 0), metric.get("mean_error", 0), metric.get("brier", 0), band, b.get("n", 0), b.get("hit_rate", 0), b.get("mean_probability", 0)]
            for col, val in enumerate(values, 1):
                cell = ws9.cell(row=reflect_row, column=col, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")
                if col in (3, 4, 5, 6, 9, 10):
                    cell.number_format = "0.00%"
            reflect_row += 1

    # ========== Sheet 2: 城池统计 ==========
    ws2 = wb.create_sheet("城池统计")
    ws2.cell(row=1, column=1, value=f"城池统计 ({date_str})").font = title_font

    row = 3
    headers = ["城池", "攻击次数", "实际概率", "理论概率", "偏差"]
    for col, h in enumerate(headers, 1):
        cell = ws2.cell(row=row, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    row += 1

    if summary and "city_stats" in summary:
        for cid in CITY_IDS:
            cs = summary["city_stats"].get(cid, {})
            name = CITY_MAP[cid]
            attacks = cs.get("total_attacks", 0)
            rate = cs.get("rate", 0)
            theo = cs.get("theoretical_prob", 0)
            bias = rate - theo
            row_data = [name, attacks, rate, theo, bias]
            for col, val in enumerate(row_data, 1):
                cell = ws2.cell(row=row, column=col, value=val)
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
                if col in [3, 4]:
                    cell.number_format = '0.00%'
                if col == 5:
                    cell.number_format = '+0.00%;-0.00%'
            row += 1

        # 卡方检验
        row += 1
        chi2 = summary.get("chi_square", {})
        ws2.cell(row=row, column=1, value="卡方检验").font = Font(bold=True)
        ws2.cell(row=row, column=2, value=f"χ²={chi2.get('chi2', 0):.2f}, p={chi2.get('p_value', 0):.4f}")
        row += 1

    # ========== Sheet 3: 间隔统计 ==========
    ws3 = wb.create_sheet("间隔统计")
    ws3.cell(row=1, column=1, value=f"上三城间隔统计 ({date_str})").font = title_font

    row = 3
    headers = ["城池", "出现次数", "平均间隔", "最小间隔", "最大间隔", "中位间隔"]
    for col, h in enumerate(headers, 1):
        cell = ws3.cell(row=row, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    row += 1

    if summary and "top3_gap_stats" in summary:
        for cid in TOP3_IDS:
            gs = summary["top3_gap_stats"].get(cid, {})
            row_data = [gs.get("name", ""), gs.get("count", 0), gs.get("mean", 0), gs.get("min", 0), gs.get("max", 0), gs.get("median", 0)]
            for col, val in enumerate(row_data, 1):
                cell = ws3.cell(row=row, column=col, value=val)
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
                if col in [3, 4, 5, 6]:
                    cell.number_format = '0.0'
            row += 1

        # 任一上三城
        any_gs = summary.get("any_top3_gap_stats", {})
        if any_gs:
            row += 1
            ws3.cell(row=row, column=1, value="任一上三城").font = Font(bold=True)
            ws3.cell(row=row, column=2, value=any_gs.get("count", 0))
            ws3.cell(row=row, column=3, value=round(any_gs.get("mean", 0), 1))
            ws3.cell(row=row, column=4, value=any_gs.get("min", 0))
            ws3.cell(row=row, column=5, value=any_gs.get("max", 0))
            ws3.cell(row=row, column=6, value=any_gs.get("median", 0))

    # ========== Sheet 4: 条件概率 ==========
    ws4 = wb.create_sheet("条件概率")
    ws4.cell(row=1, column=1, value=f"按间隔分段的条件概率 ({date_str})").font = title_font

    row = 3
    headers = ["间隔段", "样本数", "上三城命中", "上三城概率", "洛阳概率", "成都概率", "建业概率"]
    for col, h in enumerate(headers, 1):
        cell = ws4.cell(row=row, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    row += 1

    if cond_probs:
        for label in GAP_BIN_LABELS:
            cp = cond_probs.get(label, {})
            by_city = cp.get("by_city", {})
            row_data = [
                label,
                cp.get("total", 0),
                cp.get("top3_hit", 0),
                cp.get("top3_rate", 0),
                by_city.get(1, {}).get("rate", 0) if 1 in by_city else 0,
                by_city.get(2, {}).get("rate", 0) if 2 in by_city else 0,
                by_city.get(3, {}).get("rate", 0) if 3 in by_city else 0,
            ]
            for col, val in enumerate(row_data, 1):
                cell = ws4.cell(row=row, column=col, value=val)
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
                if col in [4, 5, 6, 7]:
                    cell.number_format = '0.00%'
            row += 1

    # ========== Sheet 5: 特征分析 ==========
    ws5 = wb.create_sheet("特征分析")
    ws5.cell(row=1, column=1, value=f"融合模型特征重要性 ({date_str})").font = title_font

    row = 3
    ws5.cell(row=row, column=1, value="经验间隔模板：预测概率由相似历史间隔核加权、全量命中基准率平滑，并与融合模型按回测模板融合。")
    row += 2
    if model_results:
        for cat in FUSION_CATEGORIES:
            mr = model_results.get(cat)
            if mr is None:
                continue

            ws5.cell(row=row, column=1, value=f"类别: {cat}").font = subtitle_font
            row += 1

            headers = ["特征", "显示名", "系数", "|系数|"]
            for col, h in enumerate(headers, 1):
                cell = ws5.cell(row=row, column=col, value=h)
                cell.font = header_font_white
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
            row += 1

            fi = mr.get("feature_importance", {})
            sorted_features = sorted(fi.items(), key=lambda x: x[1]["abs_coefficient"], reverse=True)
            for fn, info in sorted_features:
                row_data = [fn, info.get("display_name", fn), info.get("coefficient", 0), info.get("abs_coefficient", 0)]
                for col, val in enumerate(row_data, 1):
                    cell = ws5.cell(row=row, column=col, value=val)
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = thin_border
                    if col in [3, 4]:
                        cell.number_format = '0.0000'
                row += 1

            # 模型信息
            row += 1
            ws5.cell(row=row, column=1, value=f"样本数: {mr.get('n_samples', 0)}, 正例: {mr.get('n_positive', 0)}, 负例: {mr.get('n_negative', 0)}, 训练准确率: {mr.get('train_accuracy', 0):.2%}")
            row += 2

    # 当前经验间隔诊断，确保特征分析不是只有逻辑回归系数。
    ws5.cell(row=row, column=1, value="当前预测的经验间隔特征").font = subtitle_font
    row += 1
    exp_headers = ["目标", "当前间隔", "经验概率", "相似样本数", "当前间隔分位", "历史P25", "历史中位", "历史P75", "历史P90"]
    for col, h in enumerate(exp_headers, 1):
        cell = ws5.cell(row=row, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
    row += 1
    for key in ["any_top3"] + [TOP3_NAMES[cid] for cid in TOP3_IDS]:
        pred = (predictions or {}).get(key)
        if not pred:
            continue
        d = pred.get("experience_diagnostics", {})
        values = [pred.get("name", key), pred.get("gap", 0), pred.get("experience_probability"),
                  d.get("similar_samples", 0), d.get("gap_percentile", 0), d.get("historical_gap_p25", 0),
                  d.get("historical_gap_median", 0), d.get("historical_gap_p75", 0), d.get("historical_gap_p90", 0)]
        for col, val in enumerate(values, 1):
            cell = ws5.cell(row=row, column=col, value=val)
            cell.border = thin_border
            if col in [3, 5]:
                cell.number_format = '0.00%'
            elif col in [6, 7, 8, 9]:
                cell.number_format = '0.0'
        row += 1

    # ========== Sheet 6: 时段分析 ==========
    ws6 = wb.create_sheet("时段分析")
    ws6.cell(row=1, column=1, value=f"分时段上三城命中率 ({date_str})").font = title_font

    row = 3
    if summary and "daily_stats" in summary:
        # 按小时统计
        hourly_stats = defaultdict(lambda: {"total": 0, "top3_hit": 0, "by_city": defaultdict(int)})
        sorted_records_all = sorted(load_csv(CSV_PATH), key=lambda r: (r[0], r[1]))
        for date_str, time_str, cid, cname in sorted_records_all:
            hour = parse_hour(time_str)
            hourly_stats[hour]["total"] += 1
            if cid in TOP3_IDS:
                hourly_stats[hour]["top3_hit"] += 1
                hourly_stats[hour]["by_city"][cid] += 1

        headers = ["小时", "总手数", "上三城命中", "上三城概率", "洛阳", "成都", "建业"]
        for col, h in enumerate(headers, 1):
            cell = ws6.cell(row=row, column=col, value=h)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        row += 1

        for h in range(24):
            s = hourly_stats.get(h, {})
            total = s.get("total", 0)
            top3_hit = s.get("top3_hit", 0)
            by_city = s.get("by_city", {})
            top3_rate = top3_hit / total if total > 0 else 0
            luoyang_rate = by_city.get(1, 0) / total if total > 0 else 0
            chengdu_rate = by_city.get(2, 0) / total if total > 0 else 0
            jianye_rate = by_city.get(3, 0) / total if total > 0 else 0

            row_data = [f"{h:02d}:00", total, top3_hit, top3_rate, luoyang_rate, chengdu_rate, jianye_rate]
            for col, val in enumerate(row_data, 1):
                cell = ws6.cell(row=row, column=col, value=val)
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
                if col in [4, 5, 6, 7]:
                    cell.number_format = '0.00%'
            row += 1

    # ========== Sheet 7: 模型自检 ==========
    ws7 = wb.create_sheet("模型自检")
    ws7.cell(row=1, column=1, value=f"模型自检报告 ({date_str})").font = title_font

    row = 3
    ws7.cell(row=row, column=1, value="数据概况").font = subtitle_font
    row += 1
    total_records = summary.get("total_records", 0) if summary else 0
    ws7.cell(row=row, column=1, value="总记录数")
    ws7.cell(row=row, column=2, value=total_records)
    row += 1

    # 上三城统计
    for cid in TOP3_IDS:
        gs = summary.get("top3_gap_stats", {}).get(cid, {}) if summary else {}
        ws7.cell(row=row, column=1, value=f"{TOP3_NAMES[cid]}间隔统计")
        ws7.cell(row=row, column=2, value=f"出现{gs.get('count',0)}次, 平均间隔{gs.get('mean',0):.1f}手, 中位{gs.get('median',0)}手")
        row += 1

    row += 1
    ws7.cell(row=row, column=1, value="模型概况").font = subtitle_font
    row += 1

    if model_results:
        for cat in FUSION_CATEGORIES:
            mr = model_results.get(cat)
            if mr is None:
                ws7.cell(row=row, column=1, value=f"{cat}: 样本不足，未训练")
                row += 1
                continue
            ws7.cell(row=row, column=1, value=f"{cat}")
            ws7.cell(row=row, column=2, value=f"样本{mr.get('n_samples',0)}(正{mr.get('n_positive',0)}/负{mr.get('n_negative',0)}), 准确率{mr.get('train_accuracy',0):.2%}")
            row += 1
    else:
        ws7.cell(row=row, column=1, value="模型未训练（数据不足或sklearn不可用）")
        row += 1

    row += 1
    ws7.cell(row=row, column=1, value="严格时间序列回测（留出后段历史）").font = subtitle_font
    row += 1
    if backtest and not backtest.get("error"):
        ws7.cell(row=row, column=1, value="回测切分")
        ws7.cell(row=row, column=2, value=f"前{backtest.get('train', 0)}手训练，后{backtest.get('test', 0)}手验证")
        row += 1
        for cat in ["any_top3", "洛阳", "成都", "建业", "top1_city"]:
            metric = backtest.get(cat, {})
            label = "任一上三城" if cat == "any_top3" else cat
            if cat == "top1_city":
                text = (f"全量样本{metric.get('n', 0)}，最高概率城池命中率{metric.get('hit_rate', 0):.2%}；"
                        f"真实上三城时三城内排序命中率{metric.get('top3_conditional_hit_rate', 0):.2%}"
                        f"（{metric.get('top3_conditional_n', 0)}手）")
            else:
                text = (f"样本{metric.get('n', 0)}，阈值命中率{metric.get('hit_rate', 0):.2%}，"
                        f"真实率{metric.get('base_rate', 0):.2%}，召回率{metric.get('recall', 0):.2%}，"
                        f"Brier分数{metric.get('brier', 0):.4f}，平均概率误差{metric.get('mean_abs_error', 0):.2%}")
            ws7.cell(row=row, column=1, value=label)
            ws7.cell(row=row, column=2, value=text)
            row += 1
    else:
        ws7.cell(row=row, column=1, value=(backtest or {}).get("error", "未执行回测"))
        row += 1

    row += 1
    ws7.cell(row=row, column=1, value="建议").font = subtitle_font
    row += 1
    if total_records < 1000:
        ws7.cell(row=row, column=1, value=f"当前数据量{total_records}条，建议积累1000+条数据以获得更准确的模型")
        row += 1
    if not HAS_SKLEARN:
        ws7.cell(row=row, column=1, value="当前环境未安装sklearn，使用内置numpy逻辑回归后备模型；不是指数衰减兜底")
        row += 1

    # 设置列宽
    for ws in [ws1, ws2, ws3, ws4, ws5, ws6, ws7, ws8, ws9]:
        for col in range(1, 12):
            col_letter = chr(64 + col) if col <= 26 else 'A' + chr(64 + col - 26)
            ws.column_dimensions[col_letter].width = 14

    # 保存
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    filename = f"defense_summary_{date_str}.xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)
    temp_filepath = filepath + f".tmp.{os.getpid()}"
    wb.save(temp_filepath)
    os.replace(temp_filepath, filepath)
    return filepath


# ============================================================
# 主流程
# ============================================================
def main():
    target_date = sys.argv[1] if len(sys.argv) > 1 else "today"

    if target_date == "today":
        target_date = datetime.now().strftime("%Y-%m-%d")

    print(f"[参数] target_date={target_date}")


    try:
        ensure_dirs()

        # 加载数据
        # 加载数据：records用于模型训练（历史数据充足），tower用于间隔追踪
        records = load_csv(CSV_PATH)
        # 同时加载 tower 数据用于当前间隔
        tower_records = load_csv(TOWER_CSV_PATH) if os.path.exists(TOWER_CSV_PATH) else []
        print(f"加载{len(records)}条记录")

        if not records:
            print("ERROR: 无数据记录，请先运行采集脚本 collect_defense_data.py")
            sys.exit(1)

        # Part1: 每日汇总
        print("Part1: 计算每日汇总...")
        summary = compute_daily_summary(records, target_date)

        # Part2: 条件概率
        print("Part2: 计算条件概率...")
        cond_probs = compute_conditional_probabilities(records)

        # Part2b: 融合模型训练
        print("Part2b: 训练融合模型...")
        model_results = {}
        for cat in FUSION_CATEGORIES:
            # 先尝试弹性30维模型
            mr = train_fusion_model(records, category=cat, feature_names=ELASTIC_FEATURE_NAMES,
                                     min_positive=ELASTIC_MIN_POSITIVE, min_total=ELASTIC_MIN_TOTAL, l2_c=ELASTIC_L2_C)
            if mr is None:
                # 降级到6维模型
                mr = train_fusion_model(records, category=cat, feature_names=FUSION_FEATURE_NAMES,
                                         min_positive=FUSION_MIN_POSITIVE, min_total=FUSION_MIN_TOTAL, l2_c=FUSION_L2_C)
            model_results[cat] = mr
            if mr:
                print(f"  {cat}: 样本{mr['n_samples']}(正{mr['n_positive']}/负{mr['n_negative']}), 准确率{mr['train_accuracy']:.2%}, 特征数{len(mr['feature_names'])}")
            else:
                print(f"  {cat}: 样本不足，未训练")

        print("Part2c: 严格时间序列回测...")
        backtest = run_time_series_backtest(records)
        if backtest.get("error"):
            print(f"  回测失败: {backtest['error']}")
        else:
            for cat in ["any_top3", "洛阳", "成都", "建业"]:
                metric = backtest.get(cat, {})
                print(f"  {cat}: 验证{metric.get('n', 0)}手, 命中率{metric.get('hit_rate', 0):.2%}, Brier={metric.get('brier', 0):.4f}")
            print(f"  最高概率城池全量命中率: {backtest.get('top1_city', {}).get('hit_rate', 0):.2%}")
            print(f"  真实上三城时三城内排序命中率: {backtest.get('top1_city', {}).get('top3_conditional_hit_rate', 0):.2%}")

        # 保存模型参数
        model_params = {}
        for cat, mr in model_results.items():
            if mr:
                model_params[cat] = {
                    "feature_names": mr["feature_names"],
                    "n_samples": mr["n_samples"],
                    "n_positive": mr["n_positive"],
                    "n_negative": mr["n_negative"],
                    "train_accuracy": mr["train_accuracy"],
                    "intercept": mr["intercept"],
                    "feature_importance": {fn: info["coefficient"] for fn, info in mr["feature_importance"].items()},
                }
        save_json(MODEL_PARAMS_PATH, model_params)

        # 持久化本次采用的统一自更新模板，后续分钟刷新沿用同一套口径。
        experience_template = {}
        for cat in ["any_top3"] + [TOP3_NAMES[cid] for cid in TOP3_IDS]:
            em = build_interval_experience_model(records, cat)
            bm = build_burst_pattern_model(records, cat)
            if em:
                experience_template[cat] = {
                    "n_samples": em["n_samples"],
                    "n_positive": em["n_positive"],
                    "base_rate": em["base_rate"],
                    "completed_intervals": em["completed_intervals"],
                    "gap_mean": em["gap_mean"],
                    "gap_median": em["gap_median"],
                    "gap_p25": em["gap_p25"],
                    "gap_p75": em["gap_p75"],
                    "gap_p90": em["gap_p90"],
                    "burst_pattern": bm,
                }
        save_json(MODEL_UPDATE_TEMPLATE_PATH, {
            "version": MODEL_UPDATE_TEMPLATE_VERSION,
            "updated_at": datetime.now().isoformat(timespec="seconds"),
            "data_source": "all_records_before_current_prediction",
            "learning_template": "interval_similarity_experience_plus_fusion_blend_then_time_series_calibration",
            "experience_blend_weight": EXPERIENCE_BLEND_WEIGHT,
            "experience_prior_strength": EXPERIENCE_PRIOR_STRENGTH,
            "models": experience_template,
        })

        # 先用最新一手真实结果结算上一轮预测，使飞轮权重立即参与本轮计算。
        resolve_pending(records)

        # Part3: 概率预测
        print("Part3: 计算当前预测...")
        gap_counter = load_json(GAP_COUNTER_PATH)
        predictions = compute_current_predictions(records, model_results, gap_counter)

        # Part4: 用严格时间留出回测样本拟合校准器并应用
        print("Part4: 概率校准...")
        calibration_model = fit_probability_calibrators(backtest)
        predictions = calibrate_predictions(predictions, calibration_model)

        # 保存本次预测，等待下一手真实记录结算。
        flywheel_summary = record_snapshot(records, predictions)
        template_state = load_json(MODEL_UPDATE_TEMPLATE_PATH, {})
        template_state["flywheel_resolved_samples"] = flywheel_summary.get("resolved_samples", 0)
        template_state["flywheel_weights"] = flywheel_summary.get("weights", {})
        save_json(MODEL_UPDATE_TEMPLATE_PATH, template_state)
        flywheel_items = []
        if os.path.exists(FLYWHEEL_LOG_PATH):
            with open(FLYWHEEL_LOG_PATH, "r", encoding="utf-8") as f:
                for line in f:
                    try:
                        flywheel_items.append(json.loads(line))
                    except json.JSONDecodeError:
                        pass

        # 先发送预警，再写入 Excel；通知不再等待完整报告生成。
        print("Part5: 极高预警通知...")
        alert_result = notify_extreme_predictions(predictions)
        if alert_result.get("sent"):
            names = "、".join(item["name"] for item in alert_result.get("items", []))
            print(f"  已发送双通道通知: {names}")
        else:
            print(f"  未发送通知: {alert_result.get('reason', 'no_alert')}")

        # Part6: 生成Excel
        print("Part6: 生成Excel报告...")
        filepath = generate_excel_report(summary, cond_probs, predictions, model_results, target_date, backtest, flywheel_summary, flywheel_items)
        print(f"报告已保存: {filepath}")

        # 构建消息摘要
        msg_lines = [f"📊 斗鱼大话三国日报 ({target_date})"]
        msg_lines.append(f"数据量: {len(records)}条记录")

        if predictions:
            msg_lines.append("\n🔮 上三城预测:")
            for key in ["any_top3"] + [TOP3_NAMES[cid] for cid in TOP3_IDS]:
                if key in predictions:
                    pred = predictions[key]
                    name = pred.get("name", key)
                    gap = pred.get("gap", 0)
                    prob = pred.get("probability_calibrated", pred.get("probability", 0))
                    level = get_prob_level(prob) if 'get_prob_level' in dir() else ""
                    # 内联概率等级
                    if prob < 0.02: level = "极低"
                    elif prob < 0.05: level = "低"
                    elif prob < 0.10: level = "中"
                    elif prob < 0.20: level = "中高"
                    elif prob < 0.35: level = "高"
                    else: level = "极高"
                    msg_lines.append(f"  {name}: 间隔{gap}手, 概率{prob:.1%}({level})")

        if summary and "top3_gap_stats" in summary:
            msg_lines.append("\n📈 间隔统计:")
            for cid in TOP3_IDS:
                gs = summary["top3_gap_stats"].get(cid, {})
                msg_lines.append(f"  {TOP3_NAMES[cid]}: 平均{gs.get('mean',0):.1f}手, 中位{gs.get('median',0)}手")

        abs_path = os.path.abspath(filepath)
        msg_lines.append(f"\n📎 完整报告: {abs_path}")

        message = "\n".join(msg_lines)

        # 打印摘要到 stdout
        print(message)
        print(f"\n报告路径: {os.path.abspath(filepath)}")

    except Exception as e:
        import traceback
        traceback.print_exc()
        print(f"ERROR: 执行失败: {e}")
        sys.exit(1)


main()
